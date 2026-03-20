#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""easydsd v0.02 - DART 감사보고서 변환 도구 + Gemini AI"""

import os, re, sys, io, zipfile, threading, webbrowser, socket, time, json

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

IS_FROZEN = getattr(sys, 'frozen', False) or hasattr(sys, '_MEIPASS')
BASE_DIR  = os.path.dirname(sys.executable if IS_FROZEN else os.path.abspath(__file__))

try:
    from flask import Flask, request, send_file, jsonify, render_template_string, Response
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    if IS_FROZEN:
        print("[ERROR] 필수 라이브러리 누락. EXE를 다시 빌드하세요.")
        sys.exit(1)
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'flask', 'openpyxl', '-q'])
    from flask import Flask, request, send_file, jsonify, render_template_string, Response
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

import google.generativeai as genai  # noqa: E402

# ── 상수 ─────────────────────────────────────────────────────────────────────
def find_free_port(start=5000, end=5099):
    for p in range(start, end):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            if s.connect_ex(('127.0.0.1', p)) != 0: return p
    return start

PORT       = find_free_port()
EDIT_COLOR = 'FFF2CC'
# 회계 서식: 양수=천단위 쉼표, 음수=괄호, 0='-' 표시
FMT_ACCOUNT  = '#,##0;(#,##0);"-"'                          # 양수=1,234  음수=(1,234)  0=-
FMT_RATE     = '0.00%'                                        # 비율
FMT_DECIMAL  = '#,##0.00'                                     # 소수
SUM_COLOR  = 'E0F7FA'   # 합계/총계 행 색상
C = {'navy':'1F4E79','blue':'2E75B6','lblue':'DEEAF1',
     'yellow':'FFF2CC','white':'FFFFFF','lgray':'F2F2F2','orange':'C55A11'}
FIN_TABLE_MAP = [
    (['재 무 상 태 표'],       '🏦재무상태표'),
    (['포 괄 손 익 계 산 서'], '💹포괄손익계산서'),
    (['자 본 변 동 표'],       '📈자본변동표'),
    (['현 금 흐 름 표'],       '💰현금흐름표'),
]
FIN_PREFIXES = ('🏦','💹','📈','💰')
SUM_KEYWORDS = ['합계','총계','합 계','총 계']
PARA_COLOR   = 'E8F5E9'   # 연초록 — 주석 P/TITLE 단락 텍스트 셀

def fill(c): return PatternFill('solid', fgColor=c)
def fnt(color='000000',bold=False,size=9,italic=False):
    return Font(color=color,bold=bold,size=size,italic=italic)
def aln(h='left',v='center',wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)

# ── XML 파싱 ──────────────────────────────────────────────────────────────────
def clean_cr(s, nl=False):
    r = '\n' if nl else ' '
    s = s.replace('&amp;cr;',r).replace('&cr;',r)
    return re.sub(r'\s+',' ',s).strip()

def clean_title(s):
    s=clean_cr(s,False)
    s=s.replace('&amp;','&').replace('&lt;','<').replace('&gt;','>').replace('&quot;','"')
    return re.sub(r'\s+',' ',s).strip()

def is_blank_title(s):
    return len(re.sub(r'[&;a-z]+','',s).strip())==0

def parse_cell(m):
    attrs=m.group(1)
    val=re.sub(r'<[^>]+>','',m.group(0))
    val=(val.replace('&amp;cr;','\n').replace('&amp;','&')
           .replace('&lt;','<').replace('&gt;','>').replace('&quot;','"')
           .replace('&cr;','\n').strip())
    cs=int(x.group(1)) if (x:=re.search(r'COLSPAN="(\d+)"',attrs)) else 1
    tag=re.match(r'<([A-Z]+)',m.group(0)).group(1)
    return dict(value=val,colspan=cs,tag=tag)

def is_num_or_decimal(val):
    v=(val.strip().replace(',','').replace('(','').replace(')','')
         .replace('%','').replace('-','').replace(' ','').split('\n')[0])
    if not v: return False
    try: float(v); return True
    except: return False

def parse_xml(xml):
    exts=re.findall(r'<EXTRACTION[^>]*ACODE="([^"]+)"[^>]*>([^<]+)</EXTRACTION>',xml)
    tables=[]
    for ti,tm in enumerate(re.finditer(r'<TABLE([^>]*)>(.*?)</TABLE>',xml,re.DOTALL)):
        ctx=xml[max(0,tm.start()-600):tm.start()]
        tbody=tm.group(0)
        fin_label=next((lbl for kws,lbl in FIN_TABLE_MAP
                        if any(kw in ctx or kw in tbody for kw in kws)),'')
        raw=re.findall(r'<(?:TITLE|P)[^>]*>([^<]{3,80})</(?:TITLE|P)>',ctx)
        ctx_titles=[clean_title(t) for t in raw if not is_blank_title(t) and len(clean_title(t))>1]
        rows=[]
        for tr in re.finditer(r'<TR[^>]*>(.*?)</TR>',tm.group(2),re.DOTALL):
            cells=[parse_cell(cm) for cm in re.finditer(
                r'<(?:TD|TH|TU|TE)([^>]*)>.*?</(?:TD|TH|TU|TE)>',tr.group(1),re.DOTALL)]
            if cells: rows.append(cells)
        tables.append(dict(idx=ti,fin_label=fin_label,
                           ctx_title=(ctx_titles[-1] if ctx_titles else ''),
                           rows=rows,start=tm.start()))
    return exts,tables


def parse_paras(xml):
    """
    DSD XML에서 <P>/<TITLE> 단락 태그를 파싱.
    반환: [{'xml_start', 'xml_end', 'text', 'item_type':'P'}, ...]
    """
    paras=[]
    for m in re.finditer(r'<(?:P|TITLE)[^>]*>([^<]{5,800})</(?:P|TITLE)>',xml):
        raw=m.group(1)
        text=clean_title(raw)
        if not text or is_blank_title(raw) or len(text)<=3: continue
        paras.append({
            'item_type': 'P',
            'xml_start': m.start(),
            'xml_end':   m.end(),
            'text':      text,
        })
    return paras


def assign_paras_to_notes(paras, anchors, tables):
    """
    P단락을 주석 번호에 할당 (XML 위치 기준).
    반환: {para_idx: note_num}
    """
    if not anchors: return {}
    table_pos={t['idx']:t['start'] for t in tables}
    # 앵커: (xml_pos, note_num) 정렬
    pts=sorted([(table_pos.get(ti,0),n) for n,title,ti in anchors])
    result={}
    for pi,para in enumerate(paras):
        ppos=para['xml_start']; last_n=None
        for apos,an in pts:
            if apos<=ppos: last_n=an
            else: break
        if last_n: result[pi]=last_n
    return result

# ── 셀 헬퍼 ──────────────────────────────────────────────────────────────────
def is_edit(cell):
    f=cell.fill
    if f and f.fill_type=='solid':
        fg=f.fgColor
        if fg and fg.type=='rgb': return fg.rgb.upper().endswith(EDIT_COLOR.upper())
    return False

def is_para(cell):
    """P태그로 생성된 연초록 셀 판별"""
    f=cell.fill
    if f and f.fill_type=='solid':
        fg=f.fgColor
        if fg and fg.type=='rgb': return fg.rgb.upper().endswith(PARA_COLOR.upper())
    return False

def cell_num(v):
    if v is None: return None
    s=str(v).strip().replace(',','').replace('(','').replace(')','').replace('-','').replace(' ','')
    if not s: return None
    try:
        neg=(str(v).strip().startswith('(') and str(v).strip().endswith(')')) or str(v).strip().startswith('-')
        n=float(s)
        return -n if neg else n
    except: return None

def _to_cell_value(v):
    """
    DSD 텍스트 값을 Excel 숫자 타입으로 변환.
    → 이렇게 해야 =SUM() 수식이 실제로 합산됨

    변환 규칙:
    - '24,942,490,809' → int(24942490809)
    - '(25,608,081)'   → int(-25608081)  (음수)
    - '1.97'           → float(1.97)
    - '9', '10', '34'  → int(9), int(10) (주석번호도 무방: 롤오버 4자리 필터로 이미 제외)

    제외(문자열 유지):
    - '5,32,33' 같은 쉼표 구분 다중 번호 → 붙이면 '53233'이 되어 오작동
    - '-', 빈문자열, 변환 불가 텍스트
    """
    s = str(v).strip()
    if not s or s in ('-', ''): return v

    # 쉼표로 구분된 1~2자리 숫자 조합만 제외: '5,32,33', '6,18,32,33' 등
    # (붙이면 전혀 다른 숫자가 되므로)
    parts = s.split(',')
    if (len(parts) >= 2 and
            all(p.strip().isdigit() and 1 <= len(p.strip()) <= 2 for p in parts)):
        return v

    neg = s.startswith('(') and s.endswith(')')
    clean = s.replace(',','').replace('(','').replace(')','').replace(' ','')
    if not clean: return v
    try:
        n = float(clean) if '.' in clean else int(clean)
        return -n if neg else n
    except (ValueError, OverflowError, TypeError):
        return v


def safe_fmt(v, fmt=',.0f', fallback='없음'):
    """None/NaN/빈값 안전 포맷팅 — NoneType.__format__ 오류 방지"""
    import math as _math
    if v is None: return fallback
    try:
        fv = float(v)
        if _math.isnan(fv) or _math.isinf(fv): return fallback
        return format(fv, fmt)
    except (TypeError, ValueError):
        return fallback


# ── 기능1: 롤오버 v0.01 — 정교한 Column 지정 롤오버 ──────────────────────────
def _rollover_sheet(ws, fill_000=True):
    """
    Column-targeted 롤오버 (v0.01 버그픽스):

    [이전 버그]
    - '5,32,33' 같은 주석 참조 번호가 숫자로 오인되어
      '과목/주석 열' 데이터가 금액 열로 침범 → 데이터 파괴

    [수정 로직]
    1. 각 행에서 yellow 셀 중 '금액 셀'만 정확히 추출:
       - is_note_ref() 필터: '5,32,33' 같은 주석 참조 번호 제외
       - 4자리 이상 숫자만 인정 (소액 번호류 추가 배제)
    2. 추출된 금액 셀을 열 번호 오름차순으로 정렬
    3. [마지막-1] = 당기열,  [마지막] = 전기열  로 타겟팅
       → 좌측 과목명/주석번호 열은 1픽셀도 건드리지 않음
    4. 당기값 → 전기열 복사, 당기열 = "000" 채우기
    """
    for rowi in range(1, ws.max_row+1):
        amt_cells = []          # (col_idx, cell) — 순수 금액 셀만
        for ci in range(1, ws.max_column+1):
            cell = ws.cell(rowi, ci)
            if not is_edit(cell) or cell.value is None:
                continue
            raw = str(cell.value).strip()

            # ── 주석 참조 번호 완전 배제 ('5,32,33' 등) ──────────────
            if is_note_ref(raw):
                continue

            # ── 콤마/괄호/부호 제거 후 순수 숫자 + 4자리 이상 ───────
            vclean = (raw.replace(',','').replace('(','')
                         .replace(')','').replace('-','').replace(' ',''))
            if vclean and vclean.replace('.','').isdigit() and len(vclean) >= 4:
                amt_cells.append((ci, cell))

        # ── '-' 당기 행 감지 (정확한 판별) ─────────────────────────
        # 판별 기준: '-' 셀이 금액 셀보다 왼쪽에 있으면 당기없음 행 → skip
        # 예A) 차입금:      col3='-'(당기없음) / col5=10,000,000,000(전기)
        #       → '-'이 amt_cells 최솟값(col5)보다 왼쪽  → skip
        # 예B) 기타비유동:  col3=금액(당기) / col5='-'(전기자리)
        #       → '-'이 amt_cells 최솟값(col3)보다 오른쪽 → 단일셀 처리 진행
        if len(amt_cells) >= 1:
            min_amt_col = amt_cells[0][0]
            for ci in range(1, ws.max_column+1):
                cell_d = ws.cell(rowi, ci)
                if is_edit(cell_d) and str(cell_d.value or '').strip() == '-':
                    if ci < min_amt_col:
                        # '-' 이 금액보다 왼쪽 = 당기=0('-'), 전기=금액
                        # 롤오버: 전기 자리에 '-'(0) 이월, 당기('-') → '000'
                        _pc, p_cell = amt_cells[-1]   # 전기 위치 (마지막 금액 셀)
                        p_cell.value = '-'             # 전기에 당기값(0='-') 이월
                        cell_d.value  = '000'          # 당기 자리 = 000
                        amt_cells = []                 # 추가 처리 방지
                    break

        # 금액 셀 수에 따라 분기 처리
        if len(amt_cells) == 0:
            continue

        elif len(amt_cells) == 1:
            # ── 단일 금액 셀: 전기 위치가 '-' 또는 빈칸인 행 처리 ──────
            # 예) 기타비유동자산: col3=305,510,000 / col5='-'
            #     보증금의 증가:  col3=(330,000)  / col5='-'
            _cc, c_cell = amt_cells[0]
            p_cell = None
            # 1순위: '-' 값을 가진 노란셀 탐색 (명시적 전기 자리)
            for ci2 in range(_cc + 1, ws.max_column + 1):
                cand = ws.cell(rowi, ci2)
                if is_edit(cand) and str(cand.value or '').strip() == '-':
                    p_cell = cand
                    break
            # 2순위: '-' 없으면 빈 노란셀 탐색
            if p_cell is None:
                for ci2 in range(_cc + 1, ws.max_column + 1):
                    cand = ws.cell(rowi, ci2)
                    if is_edit(cand) and cand.value in (None, ''):
                        p_cell = cand
                        break
            if p_cell is None:
                continue  # 전기 자리 없으면 skip
            p_cell.value = c_cell.value
            c_cell.value = '000' if fill_000 else None
            continue  # 단일 셀 처리 완료

        # ── 열 인덱스 오름차순 정렬 후 마지막 두 열 타겟팅 ─────────
        # 뒤에서 두 번째 = 당기열,  마지막 = 전기열
        amt_cells.sort(key=lambda x: x[0])
        _cc, c_cell = amt_cells[-2]   # 당기
        _pc, p_cell = amt_cells[-1]   # 전기

        # ── 당기 → 전기 복사, 당기 = "000" ─────────────────────────
        p_cell.value = c_cell.value
        c_cell.value = "000" if fill_000 else None


def apply_rollover_smart(wb, api_key='', model_name='gemini-3-flash-preview'):
    """
    버그픽스 v0.01:
    - AI 판별 완전 제거
    - 4대 재무제표 본문 시트만 무조건 롤오버
      (시트명에 '재무상태표','손익계산서','자본변동표','현금흐름표' 포함 여부로 판단)
    - '주석' 이라는 단어가 포함된 시트는 절대 건드리지 않음
    - 당기 빈칸에 문자열 "000" 채우기
    """
    # 롤오버 적용 대상 키워드
    FIN_KEYWORDS = ('재무상태표', '손익계산서', '포괄손익', '자본변동표', '현금흐름표')
    NOTE_KEYWORD = '주석'

    for sname in wb.sheetnames:
        # 주석 시트는 절대 건드리지 않음
        if NOTE_KEYWORD in sname:
            continue
        # 4대 재무제표 본문 시트만 적용
        if any(kw in sname for kw in FIN_KEYWORDS):
            _rollover_sheet(wb[sname], fill_000=True)



# ── 주석 번호 앵커 추출 / 기계 분류 / AI 분류 / 그룹핑 / 기수 변경 ──────────

def extract_note_anchors(tables):
    """ctx_title에서 주석 대번호 앵커 탐지 → [(note_num,title,table_idx),...]"""
    PATS = [
        (r'^주석\s*(\d{1,2})\s*[.\-·]\s*(.{2,20})', 'prefix'),
        (r'^(\d{1,2})\s*\.\s*([^\d\(].{1,20})',                 'dot'),
        (r'^\((\d{1,2})\)\s*([^\d].{1,20})',                    'paren'),
        (r'^제\s*(\d{1,2})\s*[조절항]\s*(.{2,20})', 'clause'),
    ]
    anchors=[]; seen=set()
    for tbl in tables:
        ctx=tbl.get('ctx_title','').strip()
        if not ctx: continue
        for pat,_ in PATS:
            m=re.match(pat,ctx)
            if m:
                n=int(m.group(1)); title=m.group(2).split('&')[0].strip()[:15]
                if n not in seen and 1<=n<=99:
                    anchors.append((n,title,tbl['idx'])); seen.add(n)
                break
    return sorted(anchors,key=lambda x:x[2])


def classify_notes_machine(tables):
    """{table_idx: note_num} — 직전 앵커 기준 할당"""
    anchors=extract_note_anchors(tables)
    if not anchors: return {}
    points=sorted([(ti,n) for n,title,ti in anchors])
    result={}
    for tbl in tables:
        ti=tbl['idx']; last=None
        for ati,an in points:
            if ati<=ti: last=an
            else: break
        if last: result[ti]=last
    return result


def classify_notes_ai(api_key,tables,anchors,model_name='gemini-3-flash-preview'):
    """{table_idx: note_num} — 앵커 힌트+Gemini 추론"""
    if not api_key or not anchors: return {}
    try:
        genai.configure(api_key=api_key)
        mdl=genai.GenerativeModel(model_name)
        hint=chr(10).join(f'  TABLE[{ti}] -> 주석 {n}. {title}' for n,title,ti in anchors[:12])
        note_tbls=[t for t in tables if not t['fin_label']]
        tlist=chr(10).join(f'TABLE[{t["idx"]}] ctx={t["ctx_title"][:30]!r}' for t in note_tbls[:80])
        prompt=(
            '한국 DART 감사보고서 주석 TABLE 분류.'+chr(10)+chr(10)
            +'[이 파일의 주석 번호 패턴 힌트]'+chr(10)
            +hint+chr(10)+chr(10)
            +'[분류할 TABLE 목록]'+chr(10)+tlist+chr(10)+chr(10)
            +'각 TABLE이 몇 번 주석에 속하는지 분류해줘.'
            +'서문/목차/감사보고서 등 주석 번호 없는 것은 note=0.'+chr(10)+chr(10)
            +'JSON만 응답: {"assignment":[{"idx":30,"note":5}]}'
        )
        resp=mdl.generate_content(prompt,request_options={'timeout':90})
        m=re.search(r'\{.*\}',resp.text.strip(),re.DOTALL)
        if not m: return {}
        data=json.loads(m.group(0))
        result={item['idx']:item['note'] for item in data.get('assignment',[]) if item.get('note',0)>0}
        print(f'[Note AI] {len(result)}\uAC1C TABLE \uBD84\uB958')
        return result
    except Exception as e:
        print(f'[Note AI error] {e}'); return {}


def group_note_tables(remaining,note_assignment,notes_per_sheet=5):
    """{table_idx:note_num}으로 주석 시트 그룹핑 → [(sname,[tbl...],True),...]"""
    note_groups={}; unassigned=[]
    for tbl in remaining:
        n=note_assignment.get(tbl['idx'])
        # n이 0이거나 None이면 미분류 처리
        if n and isinstance(n,int) and n>0:
            note_groups.setdefault(n,[]).append(tbl)
        else: unassigned.append(tbl)
    sorted_notes=sorted(note_groups.keys())
    groups=[]
    for i in range(0,len(sorted_notes),notes_per_sheet):
        chunk=sorted_notes[i:i+notes_per_sheet]
        tbls=[]
        for n in chunk: tbls.extend(note_groups[n])
        fn,ln=chunk[0],chunk[-1]
        sname=(f'📝주석_{fn}' if fn==ln else f'📝주석_{fn}_{ln}')
        groups.append((sname[:31],tbls,True))
    if unassigned:
        for ci,st in enumerate(range(0,len(unassigned),10),1):
            groups.append((f'📝기타_{ci:02d}'[:31],unassigned[st:st+10],True))
    return groups


def _weave_paras(tbl_list, tbl_note_nums, all_paras, para_assign):
    """
    TABLE 리스트와 P단락을 XML 위치 순서대로 합쳐 반환.
    tbl_note_nums: 이 그룹에 포함될 주석 번호들 (P 필터용)
    all_paras: 전체 P단락 리스트
    para_assign: {para_idx: note_num}
    """
    # 이 그룹에 속하는 P단락 인덱스
    if tbl_note_nums:
        group_paras = [all_paras[pi] for pi, n in para_assign.items()
                       if n in tbl_note_nums and pi < len(all_paras)]
    else:
        # note_classify OFF: TABLE 위치 범위 안의 P만 포함
        if not tbl_list:
            return tbl_list
        min_pos = min(t['start'] for t in tbl_list)
        max_pos = max(t['start'] for t in tbl_list)
        group_paras = [p for p in all_paras
                       if min_pos - 3000 <= p['xml_start'] <= max_pos + 3000]

    items = []
    for tbl in tbl_list:
        items.append((tbl['start'], tbl))
    for para in group_paras:
        items.append((para['xml_start'], para))
    items.sort(key=lambda x: x[0])
    return [it for _, it in items]


def group_note_tables_with_paras(remaining, note_assign, note_paras, para_assign,
                                  notes_per_sheet=5):
    """
    TABLE + P단락을 주석 번호별로 그룹핑.
    반환: [(sname, [item,...], True), ...]
    """
    note_groups = {}; unassigned_t = []; unassigned_p = []
    for tbl in remaining:
        n = note_assign.get(tbl['idx'])
        if n and isinstance(n, int) and n > 0:
            note_groups.setdefault(n, {'tables': [], 'paras': []})['tables'].append(tbl)
        else:
            unassigned_t.append(tbl)

    for pi, para in enumerate(note_paras):
        n = para_assign.get(pi)
        if n and isinstance(n, int) and n > 0 and n in note_groups:
            note_groups[n]['paras'].append(para)
        elif n not in note_groups:
            unassigned_p.append(para)

    sorted_notes = sorted(note_groups.keys())
    groups = []
    for i in range(0, len(sorted_notes), notes_per_sheet):
        chunk = sorted_notes[i:i+notes_per_sheet]
        fn, ln = chunk[0], chunk[-1]
        sname = (f'📝주석_{fn}' if fn == ln else f'📝주석_{fn}_{ln}')
        # TABLE + P를 xml 위치 순서로 weave
        tbl_note_set = set(chunk)
        item_list = _weave_paras(
            [t for n in chunk for t in note_groups[n]['tables']],
            tbl_note_set, note_paras, para_assign
        )
        groups.append((sname[:31], item_list, True))

    if unassigned_t:
        for ci, start in enumerate(range(0, len(unassigned_t), 10), 1):
            chunk_t = unassigned_t[start:start+10]
            chunk_items = _weave_paras(chunk_t, set(), note_paras, para_assign)
            groups.append((f'📝기타_{ci:02d}'[:31], chunk_items, True))
    return groups



def apply_period_change(wb,cur_period,cur_year,start_m,start_d,end_m,end_d):
    """\uC7AC\uBB34\uC81C\uD45C \uD5E4\uB354 \uAE30\uC218/\uC5F0\uB3C4 \uC77C\uAD04 \uCE58\uD658"""
    prev_period=cur_period-1; prev_year=cur_year-1
    SKIP={'📋\uC0AC\uC6A9\uC548\uB0B4','_\uC6D0\uBCF8XML'}

    # 기존 기수 탐지
    old_cur_p=old_prev_p=None
    for sname in wb.sheetnames:
        if not any(sname.startswith(p) for p in FIN_PREFIXES): continue
        ws=wb[sname]
        for row in ws.iter_rows(max_row=10,values_only=True):
            for v in row:
                if not v or not isinstance(v,str): continue
                m=re.search(r'제\s*(\d{1,3})\s*\(당\)',v)
                if m and not old_cur_p: old_cur_p=int(m.group(1))
                m=re.search(r'제\s*(\d{1,3})\s*\(전\)',v)
                if m and not old_prev_p: old_prev_p=int(m.group(1))
        if old_cur_p: break
    if not old_cur_p or old_cur_p<=0:  old_cur_p=cur_period-1
    if not old_prev_p or old_prev_p<=0: old_prev_p=cur_period-2

    # 기존 연도 탐지
    year_in_hdr=[]
    for sname in wb.sheetnames:
        if not any(sname.startswith(p) for p in FIN_PREFIXES): continue
        ws=wb[sname]
        # '년' 접미사 필수 → 날짜(2025.01.01) 등 오탐 방지
        for row in ws.iter_rows(max_row=15,values_only=True):
            for v in row:
                if v and isinstance(v,str):
                    year_in_hdr+=[int(x) for x in re.findall(r'(20\d{2})년',v)]
        if len(set(year_in_hdr))>=2: break  # 당기+전기 두 연도 확보
    year_in_hdr=sorted(set(year_in_hdr))
    if len(year_in_hdr)>=2: old_prev_y,old_cur_y=year_in_hdr[0],year_in_hdr[1]
    elif len(year_in_hdr)==1: old_cur_y=year_in_hdr[0]; old_prev_y=old_cur_y-1
    else: old_cur_y=cur_year-1; old_prev_y=cur_year-2

    def rep(t):
        if not t or not isinstance(t,str): return t
        t=re.sub(rf'제\s*{old_cur_p}\s*\(당\)', f'제 {cur_period}(당)', t)
        t=re.sub(rf'제\s*{old_prev_p}\s*\(전\)', f'제 {prev_period}(전)', t)
        t=re.sub(rf'제\s*{old_cur_p}\s*기\b',     f'제 {cur_period}기',   t)
        t=re.sub(rf'제\s*{old_prev_p}\s*기\b',    f'제 {prev_period}기',  t)
        t=t.replace(f'{old_cur_y}년',  f'{cur_year}년')
        t=t.replace(f'{old_prev_y}년', f'{prev_year}년')
        return t

    for sname in wb.sheetnames:
        if sname in SKIP: continue
        ws=wb[sname]
        for rowi in range(1,ws.max_row+1):
            for ci in range(1,ws.max_column+1):
                cell=ws.cell(rowi,ci)
                if cell.value and isinstance(cell.value,str):
                    nv=rep(cell.value)
                    if nv!=cell.value: cell.value=nv



# ── 기능2: 합계/총계 행 SUM 수식 자동화 ─────────────────────────────────────
def apply_sum_formulas(wb):
    """
    모든 시트에서 첫 번째 열에 합계/총계 텍스트가 있는 행의
    yellow 숫자 셀을 =SUM() 수식으로 교체하고 연한 파란색 배경 적용
    """
    SUM_FILL=PatternFill('solid',fgColor=SUM_COLOR)
    SUM_FONT=Font(bold=True,size=9,color='006064')
    SKIP={'📋사용안내','_원본XML'}

    for sname in wb.sheetnames:
        if sname in SKIP: continue
        ws=wb[sname]
        for rowi in range(1,ws.max_row+1):
            first_val=str(ws.cell(rowi,1).value or '').strip()
            if not any(kw in first_val for kw in SUM_KEYWORDS): continue

            # 이 행의 각 열에서 위로 올라가며 SUM 범위 탐색
            for ci in range(2,ws.max_column+1):
                cell=ws.cell(rowi,ci)
                if not is_edit(cell): continue
                v=str(cell.value or '').strip()
                if not v: continue

                # 위로 올라가며 같은 열의 연속 yellow 숫자 셀 찾기
                end_r=rowi-1
                start_r=end_r
                while start_r>1:
                    above=ws.cell(start_r-1,ci)
                    if is_edit(above):
                        av=str(above.value or '').strip().replace(',','').replace('(','').replace(')','').replace('-','')
                        if av and av.replace('.','').isdigit():
                            start_r-=1; continue
                    break

                col_letter=get_column_letter(ci)
                if start_r<=end_r:
                    cell.value=f'=SUM({col_letter}{start_r}:{col_letter}{end_r})'
                cell.fill=SUM_FILL
                cell.font=SUM_FONT


# ── Python 수학 검증 ──────────────────────────────────────────────────────────
def python_verify(xlsx_bytes:bytes)->dict:
    wb=openpyxl.load_workbook(io.BytesIO(xlsx_bytes),data_only=True)
    errors=[]; warnings=[]; info=[]

    bs_sheet=next((wb[s] for s in wb.sheetnames if s.startswith('🏦')),None)
    if bs_sheet:
        asset_total=liab_total=equity_total=None
        for row in bs_sheet.iter_rows(values_only=True):
            vals=[str(v or '') for v in row]; row_txt=' '.join(vals)
            nums=[cell_num(v) for v in row if cell_num(v) is not None]
            if not nums: continue
            n=nums[0]
            if any(k in row_txt for k in ['자산총계','자산 총계','총자산']): asset_total=n
            elif any(k in row_txt for k in ['부채총계','부채 총계','총부채']): liab_total=n
            elif any(k in row_txt for k in ['자본총계','자본 총계','총자본']): equity_total=n
        if all(v is not None for v in [asset_total,liab_total,equity_total]):
            diff=abs(asset_total-(liab_total+equity_total))
            if diff>max(abs(asset_total)*1e-6,1):
                errors.append(f'[대차불일치] 자산총계({asset_total:,.0f}) != 부채({liab_total:,.0f})+자본({equity_total:,.0f}), 차이={diff:,.0f}')
            else:
                info.append(f'[대차일치] 자산={asset_total:,.0f} = 부채+자본')
        else:
            warnings.append('[대차검증] 자산총계/부채총계/자본총계 중 일부를 찾지 못했습니다.')

    fin_nums=[]; note_nums=[]
    for sname in wb.sheetnames:
        ws=wb[sname]
        for row in ws.iter_rows(max_row=150,values_only=True):
            for v in row:
                n=cell_num(v)
                if n is not None and abs(n)>0:
                    if any(sname.startswith(p) for p in FIN_PREFIXES): fin_nums.append(abs(n))
                    elif sname.startswith('📝'): note_nums.append(abs(n))
    if fin_nums and note_nums:
        fin_med=sorted(fin_nums)[len(fin_nums)//2]
        note_med=sorted(note_nums)[len(note_nums)//2]
        if fin_med>0 and note_med>0:
            ratio=max(fin_med,note_med)/min(fin_med,note_med)
            if ratio>5000:
                warnings.append(f'[단위불일치 의심] 재무제표 중간값({fin_med:,.0f}) vs 주석 중간값({note_med:,.0f}), 배율={ratio:.0f}배')

    fin_note_refs=set()
    for sname in wb.sheetnames:
        if not any(sname.startswith(p) for p in FIN_PREFIXES): continue
        ws=wb[sname]
        for row in ws.iter_rows(max_row=200,values_only=True):
            for v in row:
                if v is None or not isinstance(v,str): continue
                # 주석 참조 패턴만 탐지: '주석 5', '주석5,6', '(주 5)' 등
                # 단순 숫자(금액 등)는 제외
                for m in re.finditer(r'주\s*석?\s*(\d{1,2})(?:\s*[,·]\s*(\d{1,2}))*',str(v)):
                    nums=[int(x) for x in re.findall(r'\d{1,2}',m.group(0))]
                    fin_note_refs.update(n for n in nums if 1<=n<=50)
    existing_notes=set()
    for sname in wb.sheetnames:
        m=re.search(r'(\d{1,2})',sname)
        if m and sname.startswith('📝'):
            existing_notes.add(int(m.group(1)))
    if fin_note_refs and existing_notes:
        missing={n for n in fin_note_refs if n not in existing_notes and n>5}
        if missing:
            warnings.append(f'[주석 매핑 불완전] 참조 주석번호 {sorted(missing)} 에 해당하는 시트 없음')
        else:
            info.append(f'[주석 매핑] 참조 주석 {len(fin_note_refs)}개 모두 시트 존재')

    note_map={'refs':sorted(fin_note_refs),'existing':sorted(existing_notes),'missing':sorted({n for n in fin_note_refs if n not in existing_notes and n>5})}
    return {'errors':errors,'warnings':warnings,'info':info,'note_map':note_map}


# ── Gemini: AI 시트명 분류 ────────────────────────────────────────────────────
def gemini_classify_tables(api_key,tables,model_name='gemini-3-flash-preview'):
    if not api_key: return {}
    try:
        genai.configure(api_key=api_key)
        model=genai.GenerativeModel(model_name)
        summaries=[]
        for tbl in tables[:60]:
            vals=[c['value'] for row in tbl['rows'][:3] for c in row if c['value'].strip()][:6]
            summaries.append(f"TABLE[{tbl['idx']}] ctx={tbl['ctx_title']!r} fin={tbl['fin_label']!r} sample={vals}")
        prompt=(
            '한국 DART 감사보고서 TABLE 목록. 각 TABLE의 엑셀 시트명을 제안해주세요.'+chr(10)
            +"규칙: 재무상태표->'🏦재무상태표', 포괄손익->'💹포괄손익계산서', 자본변동->'📈자본변동표', 현금흐름->'💰현금흐름표',"+chr(10)
            +"주석->'📝주석_[주제3~5자]', 서문/목차->'📄서문', 31자이내"+chr(10)+chr(10)
            +f"TABLE:{chr(10)}{chr(10).join(summaries)}"+chr(10)+chr(10)
            +'JSON만 응답: {"mapping":[{"idx":0,"name":"예시"}]}'
        )
        resp=model.generate_content(prompt)
        m=re.search(r'\{.*\}',resp.text.strip(),re.DOTALL)
        if not m: return {}
        data=json.loads(m.group(0))
        return {item['idx']:item['name'] for item in data.get('mapping',[])}
    except Exception as e:
        print(f'[Gemini classify] {e}'); return {}


# ── Gemini: 강화 AI 교차 검증 ─────────────────────────────────────────────────
def gemini_verify_enhanced(api_key,fin_data,note_data,py_result,model_name='gemini-3-flash-preview',note_map_result=None):
    if not api_key: return 'Gemini API Key가 없습니다.'
    try:
        genai.configure(api_key=api_key)
        model=genai.GenerativeModel(model_name)
        fin_text=json.dumps(fin_data,ensure_ascii=False)[:6000]
        note_text=json.dumps(note_data,ensure_ascii=False)[:6000]
        py_err_text=chr(10).join(
            [f'[오류] {e}' for e in py_result.get('errors',[])]
            +[f'[경고] {w}' for w in py_result.get('warnings',[])]
            +[f'[정보] {i}' for i in py_result.get('info',[])]
        ) or '파이썬 자동검사 이상 없음'
        note_map_text=''
        if note_map_result and note_map_result.get('missing'):
            note_map_text=(
                chr(10)+'[주석 번호 매핑 검증]'+chr(10)
                +f'본문 참조 주석번호: {note_map_result["refs"]}'+chr(10)
                +f'실제 시트 주석번호: {note_map_result["existing"]}'+chr(10)
                +f'누락 의심 번호: {note_map_result["missing"]}'+chr(10)
                +'(누락 번호가 실제 오류인지, 다른 주석에 통합되어 있는지 확인해줘.)'+chr(10)
            )
        INST=(
            '[지시사항] 너는 공인회계사(CPA)가 아니야. '
            '복잡한 회계 기준, 적정성 여부, 감사 의견에 대한 훈수나 평가는 절대 하지 마. '
            '네 유일한 임무는 본문 재무제표의 합계 숫자와 주석에 기재된 세부 내역의 합계 숫자가 '
            '정확히 일치하는지 교차 검증(Footing)하는 것뿐이야. '
            '두 숫자가 불일치하거나 확인이 불가능한 항목만 찾아내어 간결하게 리포트해.'
        )
        prompt=(
            INST+chr(10)+chr(10)
            +f'[파이썬 1차 수학 검사 결과]{chr(10)}{py_err_text}'+chr(10)+chr(10)
            +note_map_text
            +f'[재무제표 본문]{chr(10)}{fin_text}'+chr(10)+chr(10)
            +f'[주석 데이터]{chr(10)}{note_text}'+chr(10)+chr(10)
            +'파이썬 오류와 재무제표 맥락을 합쳐서 최종 교차검증 리포트를 작성해줘.'+chr(10)+chr(10)
            +'응답 형식:'+chr(10)
            +'## ✅ 파이썬 수학 검사 결과'+chr(10)+'(자동검사 요약)'+chr(10)+chr(10)
            +'## ✅ 일치 항목'+chr(10)+'(본문과 주석 합계 일치 항목)'+chr(10)+chr(10)
            +'## ❌ 불일치 항목'+chr(10)+'(불일치 항목 + 본문금액 vs 주석합계 + 차이금액)'+chr(10)+chr(10)
            +'## ⚠️ 확인 불가 항목'+chr(10)+'(데이터 부족)'+chr(10)+chr(10)
            +'## 📋 종합'+chr(10)+'(2~3줄 요약, 회계 의견 금지)'
        )
        resp=model.generate_content(
            prompt,
            request_options={'timeout': 120}  # 120초 타임아웃
        )
        return resp.text.strip()
    except Exception as e:
        return f'Gemini API 오류: {e}'


# ── 기능3: DSD에서 당기/전기 데이터 추출 (버그3 수정: COLSPAN 확장) ─────────────
def _parse_tr_with_colspan(tr_content):
    """TR의 셀을 COLSPAN 확장하여 파싱 — 헤더와 데이터 행 열 인덱스 일치"""
    cells=[]
    for td in re.finditer(r'<(?:TD|TH|TU|TE)([^>]*)>(.*?)</(?:TD|TH|TU|TE)>',tr_content,re.DOTALL):
        attrs=td.group(1)
        v=re.sub(r'<[^>]+>','',td.group(2))
        v=v.replace('&amp;cr;',' ').replace('&amp;','&').replace('&cr;',' ').strip()
        cm=re.search(r'COLSPAN="(\d+)"',attrs,re.IGNORECASE)
        cs=int(cm.group(1)) if cm else 1
        cells.append(v)
        for _ in range(cs-1): cells.append('')  # colspan만큼 빈칸 확장
    return cells


def _find_num_in_span(row, col_start, span=2):
    """col_start 부터 span칸 내에서 첫 번째 유효한 숫자 반환 (total행/detail행 모두 대응)"""
    for offset in range(max(span, 1)):
        idx=col_start+offset
        if idx>=len(row): break
        n=cell_num(row[idx])
        if n is not None: return n
    return None


def parse_dsd_periods(dsd_bytes):
    """
    DSD 재무제표 본문 테이블에서 (fin_label, account, cur_val, pri_val) 추출.
    버그3 수정: COLSPAN 확장 파싱으로 당기/전기 열 인덱스를 정확히 타겟팅.
    """
    with zipfile.ZipFile(io.BytesIO(dsd_bytes)) as zf:
        xml=zf.read('contents.xml').decode('utf-8',errors='replace')

    results=[]
    for tm in re.finditer(r'<TABLE([^>]*)>(.*?)</TABLE>',xml,re.DOTALL):
        ctx=xml[max(0,tm.start()-600):tm.start()]
        tbody=tm.group(0)
        fin_label=next((lbl for kws,lbl in FIN_TABLE_MAP
                        if any(kw in ctx or kw in tbody for kw in kws)),'')
        if not fin_label: continue

        # COLSPAN 확장하여 모든 TR 파싱
        rows=[]
        for tr in re.finditer(r'<TR[^>]*>(.*?)</TR>',tm.group(2),re.DOTALL):
            cells=_parse_tr_with_colspan(tr.group(1))
            if cells: rows.append(cells)
        if not rows: continue

        # 헤더에서 당기/전기 열 찾기 (COLSPAN 확장 후 정확한 인덱스)
        cur_col=pri_col=None
        cur_span=pri_span=2  # 기본 span=2 (대부분 colspan=2)
        for row in rows[:6]:
            for ci,v in enumerate(row):
                if cur_col is None and '당' in v and '기' in v:
                    cur_col=ci
                    # 확장된 span 계산: 다음 셀이 빈칸이면 span=2 이상
                    span_count=1
                    while ci+span_count<len(row) and row[ci+span_count]=='':
                        span_count+=1
                    cur_span=max(span_count,1)
                elif pri_col is None and '전' in v and '기' in v:
                    pri_col=ci
                    span_count=1
                    while ci+span_count<len(row) and row[ci+span_count]=='':
                        span_count+=1
                    pri_span=max(span_count,1)
            if cur_col is not None and pri_col is not None: break

        if cur_col is None or pri_col is None: continue

        # 데이터 행에서 당기/전기 값 추출
        for row in rows:
            acct=row[0].strip() if row else ''
            if not acct or len(acct)<2: continue
            # span 범위 내에서 유효 숫자 탐색
            cur=_find_num_in_span(row, cur_col, cur_span)
            pri=_find_num_in_span(row, pri_col, pri_span)
            if cur is not None or pri is not None:
                results.append((fin_label, acct, cur, pri))
    return results

def validate_prior_period(prev_dsd,curr_dsd,api_key,model_name='gemini-3-flash-preview'):
    """
    전기금액 검증:
    prev_dsd의 당기 == curr_dsd의 전기 여야 함
    """
    prev_data=parse_dsd_periods(prev_dsd)  # (label, acct, cur, pri)
    curr_data=parse_dsd_periods(curr_dsd)

    # 전년 당기: {label: {acct: cur}}
    prev_cur={}
    for label,acct,cur,pri in prev_data:
        if cur is not None:
            prev_cur.setdefault(label,{})[acct]=cur

    # 올해 전기: {label: {acct: pri}}
    curr_pri={}
    for label,acct,cur,pri in curr_data:
        if pri is not None:
            curr_pri.setdefault(label,{})[acct]=pri

    mismatches=[]
    matches=[]
    for label,prev_d in prev_cur.items():
        curr_d=curr_pri.get(label,{})
        for acct,p_cur in prev_d.items():
            c_pri=curr_d.get(acct)
            if c_pri is None:
                mismatches.append((label,acct,p_cur,None,None))
            elif abs(p_cur-c_pri)>max(abs(p_cur)*1e-5,1):
                mismatches.append((label,acct,p_cur,c_pri,c_pri-p_cur))
            else:
                matches.append((label,acct,p_cur))

    # Gemini 리포트
    ai_report='(API Key 없음 - AI 요약 생략)'
    if api_key:
        try:
            genai.configure(api_key=api_key)
            model_obj=genai.GenerativeModel(model_name)
            mm_text=chr(10).join(
                '  ['+lb+'] '+ac+': 전기이월 오류! 수정전당기='+safe_fmt(pc)+', 수정후전기='+safe_fmt(cp)+', 차이='+safe_fmt(df,',.0f','N/A')
                for lb,ac,pc,cp,df in mismatches[:20]
            ) or '없음'
            chg_text=chr(10).join(
                '  ['+lb+'] '+ac+': 당기='+safe_fmt(cu,',.0f','-')+', 전기='+safe_fmt(pi,',.0f','-')
                for lb,ac,cu,pi in curr_data[:30]
            )
            prompt=(
                '한국 DART 감사보고서 수정전/수정후 DSD 비교 결과입니다.'+chr(10)+chr(10)
                +'[★★★ 전기이월 오류 내역 (매우 중요) ★★★]'+chr(10)
                +'수정전 DSD의 당기금액이 수정후 DSD의 전기금액과 일치하지 않는 항목:'+chr(10)
                +mm_text+chr(10)+chr(10)
                +'[수정후 DSD 재무데이터 요약]'+chr(10)+chg_text+chr(10)+chr(10)
                +'다음을 수행해:'+chr(10)
                +'1. 재무적 주요 변동 사항을 3~5줄로 요약'+chr(10)
                +'2. 특히 파이썬이 찾아낸 기초잔액 불일치 오류를 매우 강력하게 경고하는 리포트 작성'+chr(10)+chr(10)
                +'응답 형식:'+chr(10)
                +'## ⚠️ 전기이월 오류 경고'+chr(10)+'(불일치 항목 강조)'+chr(10)+chr(10)
                +'## 📊 재무적 변동 요약'+chr(10)+'(3~5줄)'+chr(10)+chr(10)
                +'## ✅ 정상 항목'+chr(10)+f'(일치 {len(matches)}건 요약)'+chr(10)+chr(10)
                +'## 📋 종합 의견'+chr(10)
            )
            resp=model_obj.generate_content(prompt)
            ai_report=resp.text.strip()
        except Exception as e:
            ai_report=f'Gemini 오류: {e}'

    return dict(mismatches=mismatches,matches=matches,
                ai_report=ai_report,curr_data=curr_data,prev_data=prev_data)


# ── DSD 비교 분석 (기존 유지) ─────────────────────────────────────────────────
def parse_dsd_tables(dsd_bytes):
    with zipfile.ZipFile(io.BytesIO(dsd_bytes)) as zf:
        xml=zf.read('contents.xml').decode('utf-8',errors='replace')
    exts=dict(re.findall(r'<EXTRACTION[^>]*ACODE="([^"]+)"[^>]*>([^<]+)</EXTRACTION>',xml))
    tables={}
    for ti,tm in enumerate(re.finditer(r'<TABLE[^>]*>(.*?)</TABLE>',xml,re.DOTALL)):
        rows=[]
        for tr in re.finditer(r'<TR[^>]*>(.*?)</TR>',tm.group(1),re.DOTALL):
            cells=[]
            for td in re.finditer(r'<(?:TD|TH|TU|TE)[^>]*>(.*?)</(?:TD|TH|TU|TE)>',tr.group(1),re.DOTALL):
                v=re.sub(r'<[^>]+>','',td.group(1))
                v=(v.replace('&amp;cr;',' ').replace('&amp;','&').replace('&cr;',' ').strip())
                cells.append(v)
            if cells: rows.append(cells)
        if rows: tables[ti]=rows
    return {'tables':tables,'exts':exts}


def compare_dsd_bytes(dsd_a,dsd_b,api_key,model_name='gemini-3-flash-preview'):
    data_a=parse_dsd_tables(dsd_a); data_b=parse_dsd_tables(dsd_b)
    ext_diffs=[]
    for k in set(list(data_a['exts'].keys())+list(data_b['exts'].keys())):
        va=data_a['exts'].get(k,'(없음)'); vb=data_b['exts'].get(k,'(없음)')
        if va!=vb: ext_diffs.append((k,va,vb))
    all_ids=sorted(set(list(data_a['tables'].keys())+list(data_b['tables'].keys())))
    table_diffs=[]
    for ti in all_ids:
        ra=data_a['tables'].get(ti,[]); rb=data_b['tables'].get(ti,[])
        for ri in range(max(len(ra),len(rb))):
            a=ra[ri] if ri<len(ra) else []; b=rb[ri] if ri<len(rb) else []
            for ci in range(max(len(a),len(b))):
                ca=a[ci] if ci<len(a) else ''; cb=b[ci] if ci<len(b) else ''
                if ca!=cb: table_diffs.append((ti,ri,ci,ca,cb))
    ai_summary='(API Key 없음)'
    if api_key:
        try:
            genai.configure(api_key=api_key)
            mdl=genai.GenerativeModel(model_name)
            dp=f'EXTRACTION변경:{len(ext_diffs)}건\nTABLE셀변경:{len(table_diffs)}건\n'
            for k,va,vb in ext_diffs[:5]: dp+=f'  {k}: {va}->{vb}\n'
            nd=[(ti,ri,ci,va,vb) for ti,ri,ci,va,vb in table_diffs if cell_num(va) or cell_num(vb)]
            for ti,ri,ci,va,vb in nd[:10]: dp+=f'  T[{ti}]R{ri}C{ci}: {va}->{vb}\n'
            prompt='한국 DART 감사보고서 두 버전 변경사항.\n'+dp+'\n주요 재무적 변동 3~5줄 요약. 증감 방향, 주요 계정 중심.'
            resp=mdl.generate_content(prompt); ai_summary=resp.text.strip()
        except Exception as e:
            ai_summary=f'Gemini오류:{e}'
    wb2=openpyxl.Workbook()
    ws0=wb2.active; ws0.title='🤖AI변동요약'; ws0.sheet_view.showGridLines=False
    tc=ws0.cell(1,1,'🔍 DSD 비교 분석 - AI 변동 요약 리포트')
    tc.fill=PatternFill('solid',fgColor='1B4F72'); tc.font=Font(color='FFFFFF',bold=True,size=13)
    tc.alignment=Alignment(horizontal='left',vertical='center')
    ws0.merge_cells('A1:D1'); ws0.row_dimensions[1].height=30
    ws0.cell(2,1,f'비교: {time.strftime("%Y-%m-%d %H:%M")}  |  EXTRACTION변경:{len(ext_diffs)}건  |  셀변경:{len(table_diffs)}건').font=Font(color='2E75B6',size=9,italic=True)
    for ri,line in enumerate(ai_summary.split('\n'),4):
        c=ws0.cell(ri,1,line); c.font=Font(size=10)
        c.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
        ws0.row_dimensions[ri].height=18
    ws0.column_dimensions['A'].width=100
    if ext_diffs:
        ws_ext=wb2.create_sheet('📊EXTRACTION변경')
        for ci,(h,w) in enumerate([('항목코드',20),('변경전',35),('변경후',35)],1):
            c=ws_ext.cell(1,ci,h); c.fill=PatternFill('solid',fgColor='1F4E79')
            c.font=Font(color='FFFFFF',bold=True,size=9); c.alignment=Alignment(horizontal='center',vertical='center')
            ws_ext.column_dimensions[get_column_letter(ci)].width=w
        for ri,(k,va,vb) in enumerate(ext_diffs,2):
            ws_ext.cell(ri,1,k).font=Font(bold=True,size=9)
            ca=ws_ext.cell(ri,2,va); cb=ws_ext.cell(ri,3,vb)
            ca.fill=PatternFill('solid',fgColor='FCE4EC'); cb.fill=PatternFill('solid',fgColor='E8F5E9')
            ca.font=Font(size=9); cb.font=Font(size=9)
    ws_d=wb2.create_sheet('🔴셀변경목록')
    for ci,(h,w) in enumerate([('TABLE','8'),('행','6'),('열','6'),('변경전','40'),('변경후','40')],1):
        c=ws_d.cell(1,ci,h); c.fill=PatternFill('solid',fgColor='1F4E79')
        c.font=Font(color='FFFFFF',bold=True,size=9); c.alignment=Alignment(horizontal='center',vertical='center')
        ws_d.column_dimensions[get_column_letter(ci)].width=int(w)*2
    RED=PatternFill('solid',fgColor='FF0000')
    for ri,(ti,rowi,ci,va,vb) in enumerate(table_diffs[:500],2):
        ws_d.cell(ri,1,ti).font=Font(size=9); ws_d.cell(ri,2,rowi).font=Font(size=9); ws_d.cell(ri,3,ci).font=Font(size=9)
        ca=ws_d.cell(ri,4,va); cb=ws_d.cell(ri,5,vb)
        ca.fill=RED; cb.fill=PatternFill('solid',fgColor='FFEB3B') if vb else RED
        ca.font=Font(size=9,color='FFFFFF'); cb.font=Font(size=9)
        ws_d.row_dimensions[ri].height=15
    if len(table_diffs)>500:
        ws_d.cell(502,1,f'(총 {len(table_diffs)}건 중 500건만 표시)').font=Font(italic=True,size=8,color='888888')
    buf=io.BytesIO(); wb2.save(buf); return buf.getvalue()


# ── 재무/주석 데이터 추출 ─────────────────────────────────────────────────────
def extract_fin_and_notes(xlsx_bytes):
    wb=openpyxl.load_workbook(io.BytesIO(xlsx_bytes),data_only=True)
    fin_data={}; note_data={}
    for sname in wb.sheetnames:
        if sname in ('📋사용안내','_원본XML','📊요약수치','🤖AI검증결과'): continue
        ws=wb[sname]
        rows_data=[]
        for row in ws.iter_rows(max_row=200,values_only=True):
            cleaned=[str(v).strip() if v is not None else '' for v in row]
            if any(cleaned): rows_data.append(cleaned)
        if any(sname.startswith(p) for p in FIN_PREFIXES): fin_data[sname]=rows_data[:80]
        elif sname.startswith('📝'): note_data[sname]=rows_data[:50]
    return fin_data,note_data


# ── DSD -> Excel 변환 (요약수치 제거, SUM 자동화 추가) ────────────────────────
def dsd_to_excel_bytes(dsd_bytes,ai_mapping=None,do_rollover=False,
                       rollover_api_key='',rollover_model='gemini-3-flash-preview',
                       do_note_classify=False,
                       do_period_change=False,period_params=None):
    with zipfile.ZipFile(io.BytesIO(dsd_bytes)) as zf:
        files={n:zf.read(n) for n in zf.namelist()}
    xml      =files.get('contents.xml',b'').decode('utf-8',errors='replace')
    meta_xml =files.get('meta.xml',b'').decode('utf-8',errors='replace')
    _exts,tables=parse_xml(xml)
    paras=parse_paras(xml)   # P/TITLE 단락 태그
    wb=openpyxl.Workbook()

    # 사용안내 (요약수치 시트 없음)
    ws0=wb.active; ws0.title='📋사용안내'; ws0.sheet_view.showGridLines=False
    guide=[
        ('DART 감사보고서 DSD - Excel 변환 도구 (easydsd v0.02)',True,C['white'],C['navy'],13),
        ('',False,'','',8),
        ('【 작업 순서 】',True,C['navy'],C['lblue'],11),
        ('  1. 노란색 셀을 당해년도 숫자/텍스트로 수정하세요',False,'000000',C['white'],10),
        ('  2. 저장 후 "Excel -> DSD" 탭에서 변환하세요',False,'000000',C['white'],10),
        ('',False,'','',8),
        ('【 색상 범례 】',True,C['navy'],C['lblue'],11),
        ('  노란색 = 수정 가능',False,'000000',C['yellow'],10),
        ('  파란색 = 헤더 (수정 불필요)',False,C['white'],C['navy'],10),
        ('  연한 파란색 = 합계/총계 (=SUM 수식 자동 삽입)',False,'006064',C[' lblue'] if '  lblue' in C else 'E0F7FA',10),
        ('',False,'','',8),
        ('【 주의사항 】',True,C['navy'],C['lblue'],11),
        ('  _원본XML 시트는 절대 수정/삭제 금지!',False,C['orange'],C['white'],10),
    ]
    # 색상 키 오타 방지
    guide[9]=(guide[9][0],guide[9][1],guide[9][2],'E0F7FA',guide[9][4])
    for ri,(txt,bold,fg,bg,sz) in enumerate(guide,1):
        cc=ws0.cell(ri,1,txt); cc.font=fnt(fg or '000000',bold=bold,size=sz)
        if bg: cc.fill=fill(bg)
        cc.alignment=aln('left',wrap=True); ws0.row_dimensions[ri].height=21
    ws0.column_dimensions['A'].width=65

    # 그룹핑
    FIN_ORDER=['🏦재무상태표','💹포괄손익계산서','📈자본변동표','💰현금흐름표']
    groups=[]; i=0
    pre_fin=[]
    while i<len(tables) and not tables[i]['fin_label']:
        pre_fin.append(tables[i]); i+=1
    if pre_fin: groups.append(('📝00_서문',pre_fin,False))
    for fin_label in FIN_ORDER:
        if i>=len(tables): break
        fin_tbls=[]
        while i<len(tables) and tables[i]['fin_label']==fin_label:
            fin_tbls.append(tables[i]); i+=1
        if not fin_tbls: continue
        if i<len(tables) and not tables[i]['fin_label']:
            fin_tbls.append(tables[i]); i+=1
        groups.append((fin_label[:31],fin_tbls,False))
    remaining=tables[i:]
    # 주석 구간 P태그: 첫 번째 note TABLE 위치 이후의 P단락
    note_sec_start=remaining[0]['start']-5000 if remaining else 0
    note_paras=[p for p in paras if p['xml_start']>=note_sec_start]
    if do_note_classify and remaining:
        anchors=extract_note_anchors(remaining)
        if rollover_api_key and anchors:
            note_assign=classify_notes_ai(rollover_api_key,remaining,anchors,rollover_model)
        else:
            note_assign=classify_notes_machine(remaining)
        para_assign=assign_paras_to_notes(note_paras,anchors,remaining) if anchors else {}
        if note_assign:
            groups.extend(group_note_tables_with_paras(
                remaining,note_assign,note_paras,para_assign,notes_per_sheet=5))
        else:
            for chunk_n,start in enumerate(range(0,len(remaining),10),1):
                chunk=_weave_paras(remaining[start:start+10],[],note_paras,{})
                groups.append((f'📝{chunk_n:02d}_주석',chunk,True))
    else:
        for chunk_n,start in enumerate(range(0,len(remaining),10),1):
            chunk=remaining[start:start+10]
            sname=f'📝{chunk_n:02d}_주석'
            if ai_mapping:
                ai_names=[ai_mapping.get(t['idx']) for t in chunk if ai_mapping.get(t['idx'])]
                sname=ai_names[0][:31] if ai_names else sname
            # P태그 weave (note_classify OFF일 때도 P태그는 포함)
            chunk=_weave_paras(chunk,[],note_paras,{})
            groups.append((sname,chunk,True))

    def write_items_to_sheet(ws,item_list,show_titles=False):
        """TABLE 아이템 + P단락 아이템을 시트에 출력. 두 딕트 반환."""
        er=1; max_cols_all=1; table_start_rows={}; para_start_rows={}
        for item in item_list:
            if item.get('item_type')=='P': continue
            tbl=item
            if not tbl['rows']: continue
            max_cols_all=max(max_cols_all,
                min(max((sum(c['colspan'] for c in row) for row in tbl['rows']),default=1),26))
        for item in item_list:
            # ── P 단락 셀 ──────────────────────────────────────────────
            if item.get('item_type')=='P':
                wc=ws.cell(er,1,item['text'])
                wc.fill=PatternFill('solid',fgColor=PARA_COLOR)
                wc.font=Font(size=9,color='1B5E20')
                wc.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
                if max_cols_all>1:
                    try: ws.merge_cells(start_row=er,start_column=1,end_row=er,end_column=min(max_cols_all,26))
                    except: pass
                n_lines=max(1,len(item['text'])//80+1)
                ws.row_dimensions[er].height=max(18,16*n_lines)
                para_start_rows[item['xml_start']]=er
                er+=1
                continue
            # ── TABLE 아이템 ───────────────────────────────────────────
            tbl=item
            if show_titles and tbl.get('ctx_title'):
                div=ws.cell(er,1,tbl['ctx_title'])
                div.fill=PatternFill('solid',fgColor='D9D9D9')
                div.font=Font(bold=True,size=9,color='333333')
                div.alignment=Alignment(horizontal='left',vertical='center')
                if max_cols_all>1:
                    try: ws.merge_cells(start_row=er,start_column=1,end_row=er,end_column=max_cols_all)
                    except: pass
                ws.row_dimensions[er].height=16; er+=1
            table_start_rows[tbl['idx']]=er
            for row in tbl['rows']:
                col=1
                for cell in row:
                    if col>26: break
                    v,tag=cell['value'],cell['tag']
                    cell_val=_to_cell_value(v) if tag not in ('TH','TE') else v
                    wc=ws.cell(er,col,cell_val)
                    if tag in ('TH','TE'):
                        wc.fill=fill(C['navy']); wc.font=fnt(C['white'],bold=True,size=9)
                        wc.alignment=aln('center',wrap=True)
                    else:
                        wc.fill=fill(C['yellow']); wc.font=fnt(size=9)
                        if isinstance(cell_val,(int,float)):
                            if isinstance(cell_val,float) and cell_val!=int(cell_val):
                                wc.number_format=FMT_DECIMAL
                            else:
                                wc.number_format=FMT_ACCOUNT
                            wc.alignment=aln('right',wrap=True)
                        else:
                            wc.alignment=aln('right' if is_num_or_decimal(v) else 'left',wrap=True)
                    if '\n' in str(v):
                        ws.row_dimensions[er].height=max(18,18*(str(v).count('\n')+1))
                    span=min(cell['colspan'],26-col+1)
                    if span>1:
                        try: ws.merge_cells(start_row=er,start_column=col,end_row=er,end_column=col+span-1)
                        except: pass
                    col+=cell['colspan']
                if not ws.row_dimensions[er].height or ws.row_dimensions[er].height<18:
                    ws.row_dimensions[er].height=18
                er+=1
        ws.column_dimensions['A'].width=28
        for ci in range(2,max_cols_all+1): ws.column_dimensions[get_column_letter(ci)].width=18
        return table_start_rows, para_start_rows

    sheet_map=[]; used=set()
    for gitem in groups:
        sraw=gitem[0]; item_list=gitem[1]; show_t=gitem[2] if len(gitem)>2 else False
        sname=sraw[:31]
        if sname in used: sname=(sraw[:28]+f'_{len(used)}')[:31]
        used.add(sname)
        ws=wb.create_sheet(sname); ws.sheet_view.showGridLines=False
        tsr,psr=write_items_to_sheet(ws,item_list,show_t)
        for item in item_list:
            if item.get('item_type')=='P':
                xstart=item['xml_start']; xend=item['xml_end']
                erow=psr.get(xstart,-1)
                sheet_map.append((sname,'P',xstart,xend,erow))
            else:
                sheet_map.append((sname,'TABLE',item['idx'],-1,tsr.get(item['idx'],-1)))

    # _원본XML
    ws_r=wb.create_sheet('_원본XML'); ws_r.sheet_view.showGridLines=False
    ws_r.cell(1,1,'이 시트는 DSD 복원에 필수입니다. 절대 수정/삭제 금지!').font=fnt(C['orange'],bold=True,size=9)
    ws_r.cell(2,1,'meta_xml'); ws_r.cell(2,2,meta_xml or '')
    HDRS=[('sheet_name',35),('idx_or_xmlstart',14),('fin_or_xmlend',22),('title_or_text',40),('excel_start_row',14),('type',8)]
    for hi,(h,w) in enumerate(HDRS,1):
        ws_r.cell(4,hi,h); ws_r.column_dimensions[get_column_letter(hi)].width=w
    for ri,(sname,rec_type,b_val,c_val,excel_row) in enumerate(sheet_map,5):
        ws_r.cell(ri,1,sname); ws_r.cell(ri,2,b_val); ws_r.cell(ri,5,excel_row)
        ws_r.cell(ri,6,rec_type)
        if rec_type=='TABLE':
            t=tables[b_val]; ws_r.cell(ri,3,t['fin_label']); ws_r.cell(ri,4,t['ctx_title'])
        else:  # P
            ws_r.cell(ri,3,c_val)  # xml_end

    # 합계/총계 SUM 수식 자동화 — 범위 오탐 문제로 비활성화
    # apply_sum_formulas(wb)

    # 롤오버
    if do_rollover:
        apply_rollover_smart(wb,api_key=rollover_api_key,model_name=rollover_model)

    # 기수/연도 자동 변경
    if do_period_change and period_params:
        try: apply_period_change(wb,**period_params)
        except Exception as pe: print(f'[period_change error] {pe}')

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


# ── Excel -> DSD 변환 ──────────────────────────────────────────────────────────
def is_note_ref(val):
    parts=val.strip().split(',')
    return (len(parts)>=2 and all(p.strip().isdigit() and 1<=len(p.strip())<=2 for p in parts))

def normalize_num(val):
    v=str(val).strip()
    if not v or v in ('-',''): return v
    if '\n' in v: return '&amp;cr;'.join(normalize_num(l) for l in v.split('\n'))
    if is_note_ref(v): return v
    neg=v.startswith('-') or (v.startswith('(') and v.endswith(')'))
    cl=v.replace(',','').replace('(','').replace(')','').replace('-','').replace(' ','')
    if cl.isdigit() and len(cl)>=3:
        fmt=f"{int(cl):,}"; v=f"({fmt})" if neg else fmt
    return v.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')

def excel_to_dsd_bytes(orig_dsd_bytes,xlsx_bytes):
    with zipfile.ZipFile(io.BytesIO(orig_dsd_bytes)) as zf:
        orig_files={n:zf.read(n) for n in zf.namelist()}
    contents_xml=orig_files['contents.xml'].decode('utf-8',errors='replace')
    wb=openpyxl.load_workbook(io.BytesIO(xlsx_bytes),data_only=True)
    mapping={}; p_records={}  # TABLE 매핑 + P태그 위치
    if '_원본XML' in wb.sheetnames:
        ws_r=wb['_원본XML']
        for row in ws_r.iter_rows(min_row=5,values_only=True):
            if not row or row[0] is None or row[1] is None: continue
            sname=str(row[0]).strip()
            rec_type=str(row[5]).strip() if len(row)>5 and row[5] else 'TABLE'
            if rec_type=='P':
                try:
                    xs=int(row[1]); xe=int(row[2]) if row[2] else 0
                    er=int(row[4]) if row[4] is not None else -1
                    if sname and xs>0: p_records.setdefault(sname,[]).append((xs,xe,er))
                except (ValueError,TypeError): pass
            else:
                try:
                    t_idx=int(row[1])
                    esr=int(row[4]) if len(row)>4 and row[4] is not None else -1
                    if sname: mapping.setdefault(sname,[]).append((t_idx,esr))
                except (ValueError,TypeError): pass
    exts={}; t_changes={}
    for sname in wb.sheetnames:
        if sname in ('📋사용안내','_원본XML','_meta'): continue
        ws=wb[sname]
        changes=[]; p_changes=[]
        for ri,row in enumerate(ws.iter_rows(min_row=2)):
            for ci,cell in enumerate(row):
                if is_edit(cell) and cell.value is not None:
                    changes.append((ri,ci,str(cell.value)))
                elif is_para(cell) and cell.value is not None:
                    p_changes.append((ri,ci,str(cell.value)))
        if changes: t_changes[sname]=changes
        if p_changes: t_changes.setdefault(sname+'__P__',[]).extend(p_changes)
    for ext_code,val in exts.items():
        contents_xml=re.sub(
            rf'(<EXTRACTION[^>]*ACODE="{re.escape(ext_code)}"[^>]*>)[^<]+(</EXTRACTION>)',
            rf'\g<1>{val}\g<2>',contents_xml)
    table_positions=[(m.start(),m.end()) for m in re.finditer(r'<TABLE[^>]*>.*?</TABLE>',contents_xml,re.DOTALL)]
    patches=[]
    for sname,changes in t_changes.items():
        t_info=mapping.get(sname)
        if not t_info: continue
        all_ch={(r,c):v for r,c,v in changes}
        for k,(t_idx,esr) in enumerate(t_info):
            if t_idx>=len(table_positions): continue
            if esr>=0:
                s=esr-2; nxt_esr=t_info[k+1][1] if k+1<len(t_info) and t_info[k+1][1]>=0 else None
                nxt=(nxt_esr-2) if nxt_esr else 99999
                local_map={(r-s,c):v for (r,c),v in all_ch.items() if s<=r<nxt}
            else:
                off=0
                for j in range(k):
                    ti2=t_info[j][0]
                    if ti2<len(table_positions):
                        sn=contents_xml[table_positions[ti2][0]:table_positions[ti2][1]]
                        off+=len(re.findall(r'<TR[^>]*>',sn))
                sn=contents_xml[table_positions[t_idx][0]:table_positions[t_idx][1]]
                tc=len(re.findall(r'<TR[^>]*>',sn))
                local_map={(r-off,c):v for (r,c),v in all_ch.items() if off<=r<off+tc}
            if not local_map: continue
            ts,te=table_positions[t_idx]; tt=contents_xml[ts:te]
            rebuilt=[]; last=0; td_row=0
            for tr_m in re.finditer(r'(<TR[^>]*>)(.*?)(</TR>)',tt,re.DOTALL):
                rebuilt.append(tt[last:tr_m.start()])
                trb=tr_m.group(2); nb=[]; td_last=0; td_col=0
                for td_m in re.finditer(r'(<(?:TD|TH|TU|TE)[^>]*>)(.*?)(</(?:TD|TH|TU|TE)>)',trb,re.DOTALL):
                    nb.append(trb[td_last:td_m.start()])
                    key=(td_row,td_col)
                    if key in local_map:
                        nb.append(td_m.group(1)+normalize_num(local_map[key])+td_m.group(3))
                    else:
                        nb.append(td_m.group(0))
                    td_last=td_m.end(); td_col+=1
                nb.append(trb[td_last:])
                rebuilt.append(tr_m.group(1)+''.join(nb)+tr_m.group(3))
                last=tr_m.end(); td_row+=1
            rebuilt.append(tt[last:])
            patches.append((ts,te,''.join(rebuilt)))
    # P태그 역변환 패치
    for sname,precs in p_records.items():
        pkey=sname+'__P__'
        pch=t_changes.get(pkey,[])
        if not pch: continue
        # {excel_row: new_text}
        row_text={ri+2:v for ri,ci,v in pch if ci==0}  # A열(ci=0), ri는 min_row=2 기준
        for xml_start,xml_end,excel_row in precs:
            new_text=row_text.get(excel_row)
            if not new_text: continue
            orig_tag=contents_xml[xml_start:xml_end]
            encoded=(new_text
                .replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
                .replace('"','&quot;').replace('\n','&amp;cr;'))
            m_tag=re.match(r'(<(?:P|TITLE)[^>]*>)(.*)(</(?:P|TITLE)>)',orig_tag,re.DOTALL)
            if m_tag:
                patches.append((xml_start,xml_end,m_tag.group(1)+encoded+m_tag.group(3)))
    result=contents_xml
    for ts,te,nt in sorted(patches,key=lambda x:-x[0]):
        result=result[:ts]+nt+result[te:]
    buf=io.BytesIO()
    with zipfile.ZipFile(buf,'w',zipfile.ZIP_DEFLATED) as zf:
        for name,data in orig_files.items():
            if os.path.splitext(name)[1].lower() in ('.jpg','.jpeg','.png','.gif','.bmp'): continue
            zf.writestr(name,result.encode('utf-8') if name=='contents.xml' else data)
    return buf.getvalue()


# ── Flask 앱 ──────────────────────────────────────────────────────────────────
app=Flask(__name__)
app.config['MAX_CONTENT_LENGTH']=100*1024*1024
_last_ping=time.time()
def _watchdog():
    time.sleep(30 if IS_FROZEN else 12)
    while True:
        time.sleep(2)
        if time.time()-_last_ping>30: os._exit(0)  # threaded 환경에서 여유 확보
threading.Thread(target=_watchdog,daemon=True).start()


# ── HTML ──────────────────────────────────────────────────────────────────────
HTML = """<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>easydsd v0.02</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Malgun Gothic',sans-serif;background:#f0f4f8;color:#1a1a2e;min-height:100vh}
.header{background:linear-gradient(135deg,#1F4E79,#2E75B6);color:white;padding:14px 24px;box-shadow:0 4px 20px rgba(31,78,121,.3)}
.hd-top{display:flex;align-items:center;justify-content:space-between;margin-bottom:10px}
.hd-top h1{font-size:17px;font-weight:700}
.hd-top p{font-size:11px;opacity:.75;margin-top:2px}
.hd-right{display:flex;align-items:center;gap:8px;flex-shrink:0}
.hd-badge{background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.3);border-radius:20px;padding:3px 12px;font-size:11px;font-weight:600}
.kill-btn{background:#c0392b;color:white;border:none;border-radius:7px;padding:6px 12px;font-size:11px;font-weight:700;cursor:pointer}
.kill-btn:hover{background:#e74c3c}
.api-bar{display:flex;align-items:center;gap:8px;background:rgba(0,0,0,.18);border-radius:8px;padding:8px 12px;flex-wrap:wrap}
.api-bar label{font-size:11px;font-weight:700;white-space:nowrap;opacity:.9}
.api-input{flex:1;min-width:160px;background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.25);border-radius:6px;color:white;padding:5px 10px;font-size:12px;font-family:monospace;outline:none}
.api-input::placeholder{opacity:.5}
.api-input:focus{background:rgba(255,255,255,.2);border-color:rgba(255,255,255,.5)}
.api-sel{background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.25);border-radius:6px;color:white;padding:4px 7px;font-size:11px;cursor:pointer;outline:none;flex-shrink:0}
.api-note{font-size:10px;white-space:nowrap;background:rgba(255,193,7,.3);border:1px solid rgba(255,193,7,.5);border-radius:10px;padding:2px 7px;color:#fff8e1;font-weight:600}
.api-st{font-size:10px;white-space:nowrap;opacity:.8}
.api-clr{background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.3);color:white;border-radius:6px;padding:3px 8px;font-size:10px;cursor:pointer;white-space:nowrap}
.api-sub{margin-top:5px;font-size:10px;opacity:.75}
.api-sub a{color:#a5d6a7;font-weight:600;text-decoration:none}
.container{max-width:860px;margin:20px auto;padding:0 16px 60px}
.tabs{display:flex;gap:2px;flex-wrap:wrap}
.tab{padding:8px 13px;border-radius:10px 10px 0 0;background:#cdd8e4;color:#4a6078;font-size:11px;font-weight:600;cursor:pointer;border:none;border-bottom:3px solid transparent;transition:all .2s;white-space:nowrap}
.tab.active{background:white;color:#1F4E79;border-bottom:3px solid #1F4E79}
.tab:hover:not(.active){background:#bcccd8}
.tab.t-ai{background:#2d1b4e;color:#b39ddb}
.tab.t-ai.active{background:white;color:#6200ea;border-bottom:3px solid #6200ea}
.tab.t-ai:hover:not(.active){background:#3d2b5e;color:#ce93d8}
.tab.t-prior{background:#1a3322;color:#80cbc4}
.tab.t-prior.active{background:white;color:#00695c;border-bottom:3px solid #00897b}
.tab.t-prior:hover:not(.active){background:#2a4a3a;color:#a5d6a7}
.tab.t-diff{background:#1a2a3a;color:#81c784}
.tab.t-diff.active{background:white;color:#1b5e20;border-bottom:3px solid #2e7d32}
.tab.t-diff:hover:not(.active){background:#2a3a4a;color:#a5d6a7}
.tab.t-dev{background:#2a2a2a;color:#999}
.tab.t-dev.active{background:white;color:#333;border-bottom:3px solid #666}
.tab.t-dev:hover:not(.active){background:#3a3a3a;color:#bbb}
.card{background:white;border-radius:0 12px 12px 12px;box-shadow:0 4px 24px rgba(0,0,0,.08);padding:24px}
.tc{display:none}.tc.active{display:block}
.step{display:flex;gap:10px;align-items:flex-start;padding:12px;margin-bottom:10px;background:#f7f9fc;border-radius:10px;border-left:4px solid #2E75B6}
.sn{min-width:24px;height:24px;border-radius:50%;background:#1F4E79;color:white;display:flex;align-items:center;justify-content:center;font-weight:700;font-size:11px;flex-shrink:0}
.st{font-weight:700;font-size:13px;color:#1F4E79;margin-bottom:3px}
.sd{font-size:11px;color:#556;line-height:1.6}
.chk{display:flex;align-items:center;gap:7px;margin-top:8px;padding:8px 11px;border-radius:8px}
.chk.teal{background:#e0f2f1;border:1px solid #80cbc4}
.chk.purple{background:#f3e5f5;border:1px solid #ce93d8}
.chk input[type=checkbox]{width:14px;height:14px;cursor:pointer}
.chk.teal input{accent-color:#00695c}
.chk.purple input{accent-color:#6200ea}
.chk label{font-size:12px;font-weight:600;cursor:pointer}
.chk.teal label{color:#004d40}
.chk.purple label{color:#4a148c}
.chk-note{font-size:10px;margin-left:4px;opacity:.7}
.dz{border:2px dashed #a0b8d0;border-radius:10px;padding:18px;text-align:center;cursor:pointer;transition:all .2s;background:#f7fbff;margin-top:7px}
.dz:hover,.dz.over{border-color:#1F4E79;background:#e8f0f8}
.dz .di{font-size:22px;margin-bottom:3px}
.dz .dl{font-size:12px;color:#4a6078}
.dz .ds{font-size:10px;color:#89a;margin-top:2px}
.fb{margin-top:5px;font-size:11px;color:#1F4E79;font-weight:600;display:none;background:#e8f0f8;padding:4px 9px;border-radius:6px}
.btn{width:100%;padding:10px;border:none;border-radius:8px;font-size:13px;font-weight:700;cursor:pointer;transition:all .2s;margin-top:10px}
.b-blue{background:linear-gradient(135deg,#1F4E79,#2E75B6);color:white}
.b-blue:hover:not(:disabled){transform:translateY(-1px);box-shadow:0 5px 18px rgba(31,78,121,.35)}
.b-blue:disabled{background:#a0b8c8;cursor:not-allowed}
.b-green{background:linear-gradient(135deg,#1a6b3a,#22a55a);color:white}
.b-green:hover:not(:disabled){transform:translateY(-1px);box-shadow:0 5px 18px rgba(26,107,58,.35)}
.b-green:disabled{background:#8fc0a0;cursor:not-allowed}
.b-ai{background:linear-gradient(135deg,#4a148c,#7b1fa2);color:white}
.b-ai:hover:not(:disabled){transform:translateY(-1px);box-shadow:0 5px 18px rgba(74,20,140,.4)}
.b-ai:disabled{background:#b39ddb;cursor:not-allowed}
.b-prior{background:linear-gradient(135deg,#00695c,#00897b);color:white}
.b-prior:hover:not(:disabled){transform:translateY(-1px);box-shadow:0 5px 18px rgba(0,105,92,.4)}
.b-prior:disabled{background:#80cbc4;cursor:not-allowed}
.b-diff{background:linear-gradient(135deg,#1b5e20,#388e3c);color:white}
.b-diff:hover:not(:disabled){transform:translateY(-1px);box-shadow:0 5px 18px rgba(27,94,32,.4)}
.b-diff:disabled{background:#81c784;cursor:not-allowed}
.pw{margin-top:10px;display:none}
.pb{height:5px;background:#e0e8f0;border-radius:4px;overflow:hidden}
.pf{height:100%;width:0%;border-radius:4px;transition:width .35s ease}
.pf-blue{background:linear-gradient(90deg,#1F4E79,#2E75B6)}
.pf-ai{background:linear-gradient(90deg,#4a148c,#7b1fa2)}
.pf-prior{background:linear-gradient(90deg,#00695c,#00897b)}
.pf-diff{background:linear-gradient(90deg,#1b5e20,#388e3c)}
.pt{font-size:11px;color:#4a6078;margin-top:3px;text-align:center}
.res{margin-top:10px;padding:11px 14px;border-radius:10px;display:none;align-items:center;gap:9px}
.res.ok{background:#e8f5ec;border:1px solid #6dbf8a}
.res.err{background:#fdecea;border:1px solid #e88}
.res.ai-ok{background:#f3e5f5;border:1px solid #ce93d8}
.res.prior-ok{background:#e0f2f1;border:1px solid #80cbc4}
.res.diff-ok{background:#e8f5e9;border:1px solid #66bb6a}
.ri{font-size:19px}
.rb{flex:1}
.rt{font-weight:700;font-size:13px}
.rs{font-size:11px;margin-top:2px;color:#556}
.dl-btn{padding:6px 11px;color:white;border:none;border-radius:6px;font-size:11px;font-weight:600;cursor:pointer;white-space:nowrap;text-decoration:none;display:inline-block;transition:background .15s}
.dl-btn.gr{background:#1a6b3a}.dl-btn.gr:hover{background:#145530}
.dl-btn.bl{background:#1F4E79}.dl-btn.bl:hover{background:#163a5e}
.dl-btn.pu{background:#4a148c}.dl-btn.pu:hover{background:#6a1b9a}
.dl-btn.te{background:#00695c}.dl-btn.te:hover{background:#004d40}
.dl-btn.fo{background:#1b5e20}.dl-btn.fo:hover{background:#145218}
.legend{display:flex;gap:9px;flex-wrap:wrap;margin-top:6px}
.li{display:flex;align-items:center;gap:5px;font-size:11px;color:#556}
.ld{width:12px;height:12px;border-radius:3px;flex-shrink:0}
.ai-notice{background:#f3e5f5;border:1px solid #ce93d8;border-radius:8px;padding:9px 13px;margin-bottom:10px;font-size:11px;color:#4a148c;line-height:1.6}
.prior-notice{background:#e0f2f1;border:1px solid #80cbc4;border-radius:8px;padding:9px 13px;margin-bottom:10px;font-size:11px;color:#004d40;line-height:1.6}
.diff-notice{background:#e8f5e9;border:1px solid #80cbc4;border-radius:8px;padding:9px 13px;margin-bottom:10px;font-size:11px;color:#1b5e20;line-height:1.6}
.sec-hdr{display:flex;align-items:center;gap:9px;padding:12px 14px;border-radius:10px;margin-bottom:12px;color:white}
.sec-hdr.purple{background:linear-gradient(135deg,#4a148c,#7b1fa2)}
.sec-hdr.teal{background:linear-gradient(135deg,#00695c,#00897b)}
.sec-hdr.diff{background:linear-gradient(135deg,#1b5e20,#2e7d32)}
.sec-hdr .ico{font-size:22px}
.sec-hdr h3{font-size:14px;font-weight:700}
.sec-hdr p{font-size:10px;opacity:.8;margin-top:2px}
.pybox{margin-top:10px;padding:10px 13px;border-radius:9px;background:#fff8e1;border:1px solid #f9a825;display:none}
.pybox h4{font-size:11px;font-weight:700;color:#f57f17;margin-bottom:5px}
.pyi{font-size:11px;padding:3px 0;border-bottom:1px solid #fff3cd;line-height:1.5}
.pyi:last-child{border-bottom:none}
.pyi.err{color:#c62828}.pyi.warn{color:#e65100}.pyi.info{color:#1b5e20}
.vbox{margin-top:10px;display:none;background:#fafafa;border:1px solid #e0e0e0;border-radius:9px;padding:12px;max-height:300px;overflow-y:auto}
.vbox h4{font-size:11px;font-weight:700;color:#4a148c;margin-bottom:5px}
.vbox pre{font-size:11px;color:#333;white-space:pre-wrap;line-height:1.7;font-family:'Malgun Gothic',sans-serif}
.two-col{display:grid;grid-template-columns:1fr 1fr;gap:10px}
.dev-pro{display:flex;align-items:center;gap:14px;padding:14px;background:linear-gradient(135deg,#1a1a2e,#16213e);border-radius:12px;margin-bottom:12px}
.dev-av{width:52px;height:52px;border-radius:50%;background:linear-gradient(135deg,#1F4E79,#2E75B6);display:flex;align-items:center;justify-content:center;font-size:22px;flex-shrink:0;border:3px solid rgba(255,255,255,.2)}
.dev-info h2{color:white;font-size:14px;font-weight:700;margin-bottom:2px}
.dev-sub{color:rgba(255,255,255,.6);font-size:10px;margin-bottom:5px}
.dev-bg{display:flex;gap:5px;flex-wrap:wrap}
.badge{border-radius:20px;padding:2px 8px;font-size:10px;font-weight:600}
.bg0{background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.2);color:rgba(255,255,255,.8)}
.bgg{background:linear-gradient(135deg,#b8860b,#daa520);color:white}
.bgt{background:rgba(46,117,182,.5);border:1px solid rgba(46,117,182,.8);color:white}
.bga{background:linear-gradient(135deg,#4a148c,#7b1fa2);color:white}
.ig{display:grid;grid-template-columns:1fr 1fr;gap:9px;margin-bottom:11px}
.ib{background:#f7f9fc;border-radius:9px;padding:10px 12px;border-left:3px solid #2E75B6}
.ib .lbl{font-size:10px;color:#89a;text-transform:uppercase;letter-spacing:.5px;margin-bottom:2px}
.ib .val{font-size:12px;font-weight:700;color:#1F4E79}
.ib .val a{color:#1F4E79;text-decoration:none}
.cbox{background:#fffbf0;border:1px solid #e8d060;border-radius:11px;padding:13px;text-align:center;margin-bottom:11px}
.ct{font-size:12px;font-weight:700;color:#7a5500;margin-bottom:5px}
.cb{font-size:12px;color:#444;line-height:1.9}
.cn{font-size:14px;font-weight:800;color:#1a1a2e;margin:4px 0 1px}
.cs{font-size:10px;color:#888}
.cc{display:inline-block;background:linear-gradient(135deg,#7c4dff,#2196f3);color:white;border-radius:20px;padding:3px 10px;font-size:11px;font-weight:700;margin:0 3px;vertical-align:middle}
.fs h3{font-size:12px;font-weight:700;color:#1F4E79;margin-bottom:6px}
.fi{display:flex;align-items:flex-start;gap:7px;padding:5px 0;border-bottom:1px solid #f0f4f8;font-size:11px;color:#446;line-height:1.5}
.fi:last-child{border-bottom:none}
.fic{font-size:12px;flex-shrink:0;margin-top:1px}
.modal-ov{display:none;position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:9999;align-items:center;justify-content:center}
.modal-ov.show{display:flex}
.modal{background:white;border-radius:13px;padding:22px 26px;max-width:300px;width:90%;text-align:center;box-shadow:0 20px 60px rgba(0,0,0,.3)}
.modal h3{font-size:14px;font-weight:700;margin-bottom:5px}
.modal p{font-size:12px;color:#556;margin-bottom:14px;line-height:1.6}
.mbtns{display:flex;gap:7px;justify-content:center}
.mbtns button{padding:8px 18px;border:none;border-radius:7px;font-size:12px;font-weight:700;cursor:pointer}
.mc{background:#e8eef4;color:#4a6078}.mc:hover{background:#d0dce8}
.mx{background:#c0392b;color:white}.mx:hover{background:#e74c3c}
.chk.orange{background:#fff3e0;border:1px solid #ffb74d}
.chk.orange input{accent-color:#e65100}.chk.orange label{color:#bf360c}
.period-row{display:none;flex-wrap:wrap;gap:6px;align-items:center;margin-top:6px;padding:8px 11px;background:#fff3e0;border-radius:7px;font-size:11px;color:#bf360c}
.pi{width:58px;padding:4px 6px;border:1px solid #ffb74d;border-radius:5px;font-size:12px;text-align:center;background:#fff8f0;outline:none}
.pi:focus{border-color:#e65100}
.period-sep{color:#999;font-size:11px}
.chk.navy{background:#e8eef4;border:1px solid #90adc4}
.chk.navy input{accent-color:#1F4E79}.chk.navy label{color:#1F4E79}
</style>
</head>
<body>
<div class="header">
  <div class="hd-top">
    <div>
      <h1>&#128202; DART 감사보고서 변환 도구</h1>
      <p>DSD &harr; Excel &nbsp;&#xB7;&nbsp; AI 검증 &nbsp;&#xB7;&nbsp; 전기금액 검증 &nbsp;&#xB7;&nbsp; 롤오버 &nbsp;&#xB7;&nbsp; easydsd v0.02</p>
    </div>
    <div class="hd-right">
      <div class="hd-badge">v0.01</div>
      <button class="kill-btn" onclick="showKill()">&#x23FC; 종료</button>
    </div>
  </div>
  <div class="api-bar">
    <label>&#128273; Gemini API Key</label>
    <input class="api-input" id="apiKey" type="password"
      placeholder="AIza... (없어도 DSD 변환/롤오버 완벽 작동)"
      oninput="saveKey(this.value)" />
    <select id="modelSel" class="api-sel" onchange="saveModel(this.value)">
      <option value="gemini-3-flash-preview">3 Flash &#x2605; 최신</option>
      <option value="gemini-2.5-flash">2.5 Flash</option>
      <option value="gemini-2.0-flash">2.0 Flash (6월종료)</option>
    </select>
    <span class="api-note">&#128204; 선택사항</span>
    <span class="api-st" id="apiSt">&#x26AA; 미입력</span>
    <button class="api-clr" onclick="clearKey()">&#x2715; 삭제</button>
  </div>
  <div class="api-sub">
    &#128204; API Key 없이도 DSD&harr;Excel 변환 및 롤오버(FIN 시트)는 완벽 작동합니다. &nbsp;|&nbsp;
    <a href="https://aistudio.google.com/app/apikey" target="_blank">&#128073; 무료 API 키 발급 &#x2197;</a>
  </div>
</div>

<div class="modal-ov" id="killModal">
  <div class="modal">
    <h3>&#x26A0;&#xFE0F; 종료할까요?</h3>
    <p>서버 프로세스가 완전히 종료됩니다.</p>
    <div class="mbtns">
      <button class="mc" onclick="hideKill()">취소</button>
      <button class="mx" onclick="doKill()">종료</button>
    </div>
  </div>
</div>

<div class="container">
  <div class="tabs">
    <button class="tab active" onclick="sw(0)">&#9312; DSD &#8594; Excel</button>
    <button class="tab t-ai" onclick="sw(1)">&#129302; AI 재무제표 검증</button>
    <button class="tab" onclick="sw(2)">&#9313; Excel &#8594; DSD</button>
    <button class="tab t-prior" onclick="sw(3)">&#128270; 전기금액 검증</button>
    <button class="tab t-diff" onclick="sw(4)">&#128269; DSD 비교분석</button>
    <button class="tab t-dev" onclick="sw(5)">개발자 정보</button>
  </div>
  <div class="card">

    <!-- 탭0: DSD->Excel -->
    <div class="tc active" id="tc0">
      <div class="step">
        <div class="sn">1</div>
        <div>
          <div class="st">전년도 DSD 파일을 업로드하세요</div>
          <div class="sd">DART에서 제출한 .dsd 파일을 드래그하거나 클릭해 선택하세요.<br>변환된 Excel의 <b>노란색 셀</b>을 당해년도 숫자로 수정하시면 됩니다.</div>
          <div class="dz" id="dz1" onclick="document.getElementById('f1').click()"
               ondragover="dov(event,'dz1')" ondragleave="dlv('dz1')" ondrop="ddrop(event,'f1','dz1')">
            <div class="di">&#128194;</div>
            <div class="dl">클릭하거나 파일을 여기에 끌어다 놓으세요</div>
            <div class="ds">.dsd 파일</div>
          </div>
          <input type="file" id="f1" accept=".dsd" style="display:none" onchange="sf('f1','fb1','dz1')">
          <div class="fb" id="fb1"></div>
          <div class="chk teal">
            <input type="checkbox" id="chkRoll">
            <label for="chkRoll">&#128260; 롤오버 (작년 당기&#8594;올해 전기 이월, 당기 칸 000 채움)</label>
          </div>
          <div class="chk orange">
            <input type="checkbox" id="chkPeriod" onchange="document.getElementById('periodRow').style.display=this.checked?'flex':'none'">
            <label for="chkPeriod">&#128197; 기수/연도 자동 변경 (헤더 텍스트 일괄 치환)</label>
          </div>
          <div class="period-row" id="periodRow">
            <label>당기</label>
            <input class="pi" id="curPeriod" type="number" min="1" max="999" placeholder="40">
            <span class="period-sep">기</span>
            <input class="pi" id="curYear" type="number" min="2000" max="2099" placeholder="2026">
            <span class="period-sep">년</span>
            <input class="pi" id="startM" type="number" min="1" max="12" placeholder="1">
            <span class="period-sep">월</span>
            <input class="pi" id="startD" type="number" min="1" max="31" placeholder="1">
            <span class="period-sep">일&#126;</span>
            <input class="pi" id="endM" type="number" min="1" max="12" placeholder="12">
            <span class="period-sep">월</span>
            <input class="pi" id="endD" type="number" min="1" max="31" placeholder="31">
            <span class="period-sep">일</span>
            <span style="color:#999;font-size:10px">(전기 자동계산)</span>
          </div>
          <div class="chk purple">
            <input type="checkbox" id="chkNote">
            <label for="chkNote">&#128218; 주석 번호별 자동 분류 (API없이 기계파싱 / API있으면 Gemini 정밀분류)</label>
          </div>
          <div class="chk navy">
            <input type="checkbox" id="chkAI">
            <label for="chkAI">&#129302; AI 스마트 분류 (주석 Gemini 활용 &#xB7; API Key &#xD544;&#xC694;)</label>
          </div>
        </div>
      </div>
      <button class="btn b-blue" id="btn1" onclick="run1()" disabled>&#128229; Excel 파일로 변환하기</button>
      <div class="pw" id="pw1"><div class="pb"><div class="pf pf-blue" id="pf1"></div></div><div class="pt" id="pt1">변환 중...</div></div>
      <div class="res ok" id="ok1">
        <div class="ri">&#9989;</div>
        <div class="rb">
          <div class="rt" id="ok1t"></div><div class="rs" id="ok1s"></div>
          <div class="legend" style="margin-top:6px">
            <div class="li"><div class="ld" style="background:#FFF2CC;border:1px solid #ccc"></div>노란색=수정가능</div>
            <div class="li"><div class="ld" style="background:#1F4E79"></div>파란색=헤더</div>
            <div class="li"><div class="ld" style="background:#E0F7FA;border:1px solid #80cbc4"></div>하늘색=합계(SUM)</div>
          </div>
        </div>
        <a class="dl-btn gr" id="dl1" href="#">&#11015; 다운로드</a>
      </div>
      <div class="res err" id="er1"><div class="ri">&#10060;</div><div class="rb"><div class="rt">변환 실패</div><div class="rs" id="er1m"></div></div></div>
    </div>

    <!-- 탭1: AI 재무제표 검증 -->
    <div class="tc" id="tc1">
      <div class="ai-notice">
        &#x2139; <b>이 기능은 선택사항입니다.</b> API Key 없이도 Python 수학 검사는 실행됩니다.<br>
        &#128270; <b>Python 자동 검사</b>(대차평균, 단위 이상, 주석 매핑)는 API Key 없이도 항상 동작합니다.
      </div>
      <div class="sec-hdr purple">
        <div class="ico">&#129302;</div>
        <div><h3>AI 재무제표 교차 검증 (강화판)</h3><p>Python 수학 검사 + Gemini AI Footing 검증</p></div>
      </div>
      <div class="step">
        <div class="sn">1</div>
        <div>
          <div class="st">수정한 Excel 파일 업로드</div>
          <div class="sd">easydsd로 변환 후 수정한 .xlsx 파일을 올려주세요.<br>
            <b style="color:#4a148c">&#x26A0; Gemini AI 검증은 API Key가 필요합니다.</b></div>
          <div class="dz" id="dz4" onclick="document.getElementById('f4').click()"
               ondragover="dov(event,'dz4')" ondragleave="dlv('dz4')" ondrop="ddrop(event,'f4','dz4')">
            <div class="di">&#128202;</div><div class="dl">수정된 Excel (.xlsx)</div><div class="ds">easydsd 변환 파일</div>
          </div>
          <input type="file" id="f4" accept=".xlsx" style="display:none" onchange="sf('f4','fb4','dz4')">
          <div class="fb" id="fb4"></div>
        </div>
      </div>
      <div class="chk navy" style="margin-bottom:6px">
        <input type="checkbox" id="chkNoteMap">
        <label for="chkNoteMap">&#128279; 주석 번호 매핑 검증 (본문 참조번호 vs 실제 주석 시트 대조, API있으면 AI 2차확인)</label>
      </div>
      <button class="btn b-ai" id="btn3" onclick="run3()" disabled>&#129302; AI 교차 검증 실행하기</button>
      <div class="pw" id="pw3"><div class="pb"><div class="pf pf-ai" id="pf3"></div></div><div class="pt" id="pt3">분석 중...</div></div>
      <div class="pybox" id="pybox"><h4>&#128270; Python 수학 검사 결과</h4><div id="pyitems"></div></div>
      <div class="res ai-ok" id="ok3">
        <div class="ri">&#129302;</div>
        <div class="rb"><div class="rt" id="ok3t"></div><div class="rs" id="ok3s"></div></div>
        <a class="dl-btn pu" id="dl3" href="#">&#11015; 검증결과 다운로드</a>
      </div>
      <div class="res err" id="er3"><div class="ri">&#10060;</div><div class="rb"><div class="rt">검증 실패</div><div class="rs" id="er3m"></div></div></div>
      <div class="vbox" id="vbox"><h4>&#129302; AI 결과 미리보기</h4><pre id="vtext"></pre></div>
    </div>

    <!-- 탭2: Excel->DSD -->
    <div class="tc" id="tc2">
      <div class="step">
        <div class="sn">1</div>
        <div>
          <div class="st">원본 DSD 파일 업로드</div>
          <div class="sd">&#9312; 탭에서 사용했던 원본 .dsd 파일을 올려주세요.</div>
          <div class="dz" id="dz2" onclick="document.getElementById('f2').click()"
               ondragover="dov(event,'dz2')" ondragleave="dlv('dz2')" ondrop="ddrop(event,'f2','dz2')">
            <div class="di">&#128194;</div><div class="dl">원본 DSD 파일</div><div class="ds">.dsd</div>
          </div>
          <input type="file" id="f2" accept=".dsd" style="display:none" onchange="sf('f2','fb2','dz2')">
          <div class="fb" id="fb2"></div>
        </div>
      </div>
      <div class="step">
        <div class="sn">2</div>
        <div>
          <div class="st">수정한 Excel 파일 업로드</div>
          <div class="sd">노란색 셀을 수정한 .xlsx 파일을 올려주세요.</div>
          <div class="dz" id="dz3" onclick="document.getElementById('f3').click()"
               ondragover="dov(event,'dz3')" ondragleave="dlv('dz3')" ondrop="ddrop(event,'f3','dz3')">
            <div class="di">&#128202;</div><div class="dl">수정된 Excel 파일</div><div class="ds">.xlsx</div>
          </div>
          <input type="file" id="f3" accept=".xlsx" style="display:none" onchange="sf('f3','fb3','dz3')">
          <div class="fb" id="fb3"></div>
        </div>
      </div>
      <button class="btn b-green" id="btn2" onclick="run2()" disabled>&#128228; DSD 파일로 변환하기</button>
      <div class="pw" id="pw2"><div class="pb"><div class="pf pf-blue" id="pf2"></div></div><div class="pt" id="pt2">변환 중...</div></div>
      <div class="res ok" id="ok2">
        <div class="ri">&#9989;</div>
        <div class="rb"><div class="rt" id="ok2t"></div><div class="rs" id="ok2s"></div></div>
        <a class="dl-btn bl" id="dl2" href="#">&#11015; DSD 다운로드</a>
      </div>
      <div class="res err" id="er2"><div class="ri">&#10060;</div><div class="rb"><div class="rt">변환 실패</div><div class="rs" id="er2m"></div></div></div>
    </div>

    <!-- 탭3: 전기금액 검증 -->
    <div class="tc" id="tc3">
      <div class="prior-notice">
        &#128270; <b>전기금액 검증</b>: 수정 전 DSD의 당기금액이 수정 후 DSD의 전기금액과 일치하는지 자동으로 확인합니다.<br>
        기초잔액 불일치(전기이월 오류)를 발견하면 AI가 강력하게 경고합니다. API Key 없이도 Python 비교는 실행됩니다.
      </div>
      <div class="sec-hdr teal">
        <div class="ico">&#128270;</div>
        <div><h3>전기금액 검증 (기초잔액 검사)</h3><p>수정전DSD 당기 == 수정후DSD 전기 여부를 항목별로 검증합니다</p></div>
      </div>
      <div class="two-col">
        <div class="step">
          <div class="sn">1</div>
          <div>
            <div class="st">수정 전 DSD (작년)</div>
            <div class="sd">작년에 제출한 원본 DSD 파일</div>
            <div class="dz" id="dz5" onclick="document.getElementById('f5').click()"
                 ondragover="dov(event,'dz5')" ondragleave="dlv('dz5')" ondrop="ddrop(event,'f5','dz5')">
              <div class="di">&#128194;</div><div class="dl">수정 전 DSD (작년)</div><div class="ds">.dsd</div>
            </div>
            <input type="file" id="f5" accept=".dsd" style="display:none" onchange="sf('f5','fb5','dz5')">
            <div class="fb" id="fb5"></div>
          </div>
        </div>
        <div class="step">
          <div class="sn">2</div>
          <div>
            <div class="st">수정 후 DSD (올해)</div>
            <div class="sd">올해 작성/수정한 DSD 파일</div>
            <div class="dz" id="dz6" onclick="document.getElementById('f6').click()"
                 ondragover="dov(event,'dz6')" ondragleave="dlv('dz6')" ondrop="ddrop(event,'f6','dz6')">
              <div class="di">&#128194;</div><div class="dl">수정 후 DSD (올해)</div><div class="ds">.dsd</div>
            </div>
            <input type="file" id="f6" accept=".dsd" style="display:none" onchange="sf('f6','fb6','dz6')">
            <div class="fb" id="fb6"></div>
          </div>
        </div>
      </div>
      <button class="btn b-prior" id="btn4" onclick="runPrior()" disabled>&#128270; 전기금액 검증 실행</button>
      <div class="pw" id="pw4"><div class="pb"><div class="pf pf-prior" id="pf4"></div></div><div class="pt" id="pt4">비교 분석 중...</div></div>
      <div class="res prior-ok" id="ok4">
        <div class="ri">&#128270;</div>
        <div class="rb"><div class="rt" id="ok4t"></div><div class="rs" id="ok4s"></div></div>
        <a class="dl-btn te" id="dl4" href="#">&#11015; 검증결과 다운로드</a>
      </div>
      <div class="res err" id="er4"><div class="ri">&#10060;</div><div class="rb"><div class="rt">검증 실패</div><div class="rs" id="er4m"></div></div></div>
    </div>

    <!-- 탭4: DSD 비교분석 -->
    <div class="tc" id="tc4">
      <div class="diff-notice">
        &#128269; <b>DSD 비교 분석</b>: 두 DSD 파일의 XML을 셀 단위로 비교하여 변경된 부분을 강조합니다.<br>
        <b style="color:#1b5e20">&#x26A0; 해당 기능은 당기 DSD 작성을 완료하고, 수정되었을 때 확인하는 용도로 사용하시기 바랍니다.</b>
      </div>
      <div class="sec-hdr diff">
        <div class="ico">&#128269;</div>
        <div><h3>DSD 비교 분석 (Diff)</h3><p>수정 전/후 DSD 파일의 재무적 변동을 자동 감지합니다</p></div>
      </div>
      <div class="two-col">
        <div class="step">
          <div class="sn">A</div>
          <div>
            <div class="st">수정 전 DSD</div>
            <div class="dz" id="dz7" onclick="document.getElementById('f7').click()"
                 ondragover="dov(event,'dz7')" ondragleave="dlv('dz7')" ondrop="ddrop(event,'f7','dz7')">
              <div class="di">&#128194;</div><div class="dl">수정 전 DSD</div><div class="ds">.dsd</div>
            </div>
            <input type="file" id="f7" accept=".dsd" style="display:none" onchange="sf('f7','fb7','dz7')">
            <div class="fb" id="fb7"></div>
          </div>
        </div>
        <div class="step">
          <div class="sn">B</div>
          <div>
            <div class="st">수정 후 DSD</div>
            <div class="dz" id="dz8" onclick="document.getElementById('f8').click()"
                 ondragover="dov(event,'dz8')" ondragleave="dlv('dz8')" ondrop="ddrop(event,'f8','dz8')">
              <div class="di">&#128194;</div><div class="dl">수정 후 DSD</div><div class="ds">.dsd</div>
            </div>
            <input type="file" id="f8" accept=".dsd" style="display:none" onchange="sf('f8','fb8','dz8')">
            <div class="fb" id="fb8"></div>
          </div>
        </div>
      </div>
      <button class="btn b-diff" id="btn5" onclick="runDiff()" disabled>&#128269; DSD 비교 분석 실행</button>
      <div class="pw" id="pw5"><div class="pb"><div class="pf pf-diff" id="pf5"></div></div><div class="pt" id="pt5">비교 중...</div></div>
      <div class="res diff-ok" id="ok5">
        <div class="ri">&#128269;</div>
        <div class="rb"><div class="rt" id="ok5t"></div><div class="rs" id="ok5s"></div></div>
        <a class="dl-btn fo" id="dl5" href="#">&#11015; Diff 리포트 다운로드</a>
      </div>
      <div class="res err" id="er5"><div class="ri">&#10060;</div><div class="rb"><div class="rt">비교 실패</div><div class="rs" id="er5m"></div></div></div>
    </div>

    <!-- 탭5: 개발자 정보 -->
    <div class="tc" id="tc5">
      <div class="dev-pro">
        <div class="dev-av">&#127970;</div>
        <div class="dev-info">
          <h2>easydsd v0.02</h2>
          <div class="dev-sub">DART 감사보고서 DSD 변환 + AI 검증 + 전기금액 검증 + DSD 비교</div>
          <div class="dev-bg">
            <span class="badge bg0">v0.01</span>
            <span class="badge bgg">&#129302; AI-Powered</span>
            <span class="badge bgt">Python+Flask</span>
            <span class="badge bga">Gemini 3 Flash</span>
          </div>
        </div>
      </div>
      <div class="ig">
        <div class="ib"><div class="lbl">개발자</div><div class="val"><a href="mailto:eeffco11@naver.com">eeffco11@naver.com</a></div></div>
        <div class="ib"><div class="lbl">버전</div><div class="val">easydsd v0.02</div></div>
        <div class="ib"><div class="lbl">지원 파일</div><div class="val">.dsd / .xlsx</div></div>
        <div class="ib"><div class="lbl">AI 엔진</div><div class="val">Gemini 3 Flash</div></div>
      </div>
      <div class="cbox">
        <div class="ct">&#128591; 제작 크레딧</div>
        <div class="cb">이 프로그램은 전적으로<br><span class="cc">Claude (Anthropic)</span>가 설계하고 개발했습니다.
          <div class="cn">클로드 짱짱맨</div><div class="cs">전 과정을 클로드로 다함</div>
        </div>
      </div>
      <div class="fs">
        <h3>&#10024; v0.01 주요 기능</h3>
        <div class="fi"><div class="fic">&#128260;</div><div><b>스마트 롤오버</b> - FIN 시트 무조건 + 주석 AI 판별, 당기 칸 000 자동 채움</div></div>
        <div class="fi"><div class="fic">&#8721;</div><div><b>SUM 수식 자동화</b> - 합계/총계 행을 =SUM() 수식으로 자동 교체 + 하늘색 강조</div></div>
        <div class="fi"><div class="fic">&#128270;</div><div><b>전기금액 검증</b> - 수정전 당기 == 수정후 전기 일치 여부 항목별 자동 검사</div></div>
        <div class="fi"><div class="fic">&#129302;</div><div><b>AI 강화 검증</b> - Python 수학 검사 + Gemini Footing 교차검증 (CPA 의견 금지)</div></div>
        <div class="fi"><div class="fic">&#128269;</div><div><b>DSD 비교분석</b> - 셀 단위 Diff + AI 변동 요약 + 빨간 하이라이트</div></div>
        <div class="fi"><div class="fic">&#128163;</div><div>하트비트 감시 - 브라우저 닫으면 서버 자동 종료, 종료시 API Key 삭제</div></div>
      </div>
    </div>

  </div>
</div>

<script>
function loadKey(){var k=localStorage.getItem('easydsd_k')||'';document.getElementById('apiKey').value=k;updSt(k);}
function saveKey(v){if(v)localStorage.setItem('easydsd_k',v);else localStorage.removeItem('easydsd_k');updSt(v);}
function clearKey(){localStorage.removeItem('easydsd_k');document.getElementById('apiKey').value='';updSt('');}
function getKey(){return localStorage.getItem('easydsd_k')||'';}
function updSt(v){var e=document.getElementById('apiSt');if(v&&v.length>10){e.textContent='\\uD83D\\uDFE2 입력됨';e.style.color='#a5d6a7';}else{e.textContent='\\u26AA 미입력';e.style.color='rgba(255,255,255,.6)';}}
function saveModel(v){localStorage.setItem('easydsd_m',v);}
function getModel(){return localStorage.getItem('easydsd_m')||'gemini-3-flash-preview';}
function loadModel(){var m=getModel();var s=document.getElementById('modelSel');if(s)s.value=m;}
loadKey();loadModel();
setInterval(function(){fetch('/api/heartbeat',{method:'POST'}).catch(function(){});},2500);
var F={f1:null,f2:null,f3:null,f4:null,f5:null,f6:null,f7:null,f8:null};
function sw(n){document.querySelectorAll('.tab').forEach(function(t,i){t.classList.toggle('active',i===n);});document.querySelectorAll('.tc').forEach(function(t,i){t.classList.toggle('active',i===n);});}
function sf(id,bid,dzId){var f=document.getElementById(id).files[0];if(!f)return;F[id]=f;var b=document.getElementById(bid);b.textContent='\\u2713  '+f.name+'  ('+(f.size/1024).toFixed(0)+' KB)';b.style.display='block';document.getElementById(dzId).style.borderColor='#1F4E79';chk();}
function dov(e,id){e.preventDefault();document.getElementById(id).classList.add('over');}
function dlv(id){document.getElementById(id).classList.remove('over');}
function ddrop(e,fid,did){e.preventDefault();dlv(did);var dt=e.dataTransfer;if(!dt.files.length)return;var inp=document.getElementById(fid);var tr=new DataTransfer();tr.items.add(dt.files[0]);inp.files=tr.files;sf(fid,fid.replace('f','fb'),did);}
function chk(){
  document.getElementById('btn1').disabled=!F.f1;
  document.getElementById('btn2').disabled=!(F.f2&&F.f3);
  document.getElementById('btn3').disabled=!F.f4;
  document.getElementById('btn4').disabled=!(F.f5&&F.f6);
  document.getElementById('btn5').disabled=!(F.f7&&F.f8);
}
function hide(n){['ok','er'].forEach(function(p){document.getElementById(p+n).style.display='none';});}
var piv=null;
function sp(n,msg,isAI){
  hide(n);
  if(n===3){document.getElementById('vbox').style.display='none';document.getElementById('pybox').style.display='none';}
  var pw=document.getElementById('pw'+n);pw.style.display='block';
  document.getElementById('pt'+n).textContent=msg;
  document.getElementById('pf'+n).style.width='0%';
  var w=0;piv=setInterval(function(){w=Math.min(w+(isAI?1:4),88);document.getElementById('pf'+n).style.width=w+'%';},isAI?400:200);
}
function ep(n){clearInterval(piv);document.getElementById('pf'+n).style.width='100%';setTimeout(function(){document.getElementById('pw'+n).style.display='none';},500);}
function sok(n,t,s,blob,fname){var b=document.getElementById('ok'+n);b.style.display='flex';document.getElementById('ok'+n+'t').textContent=t;document.getElementById('ok'+n+'s').textContent=s;if(blob){var dl=document.getElementById('dl'+n);dl.href=URL.createObjectURL(blob);dl.download=fname;}}
function ser(n,msg){var b=document.getElementById('er'+n);b.style.display='flex';document.getElementById('er'+n+'m').textContent=msg;}
function showPy(r){var box=document.getElementById('pybox');var items=document.getElementById('pyitems');items.innerHTML='';var all=(r.errors||[]).map(function(e){return{t:'err',v:e};}).concat((r.warnings||[]).map(function(w){return{t:'warn',v:w};})).concat((r.info||[]).map(function(i){return{t:'info',v:i};}));if(!all.length){items.innerHTML='<div class="pyi info">\\u2705 \\uc774\\uc0c1 \\uc5c6\\uc74c</div>';box.style.display='block';return;}all.forEach(function(item){var d=document.createElement('div');d.className='pyi '+item.t;d.textContent=(item.t==='err'?'\\u274C ':item.t==='warn'?'\\u26A0\\uFE0F ':'\\u2705 ')+item.v;items.appendChild(d);});box.style.display='block';}
var S1=['DSD \\ud30c\\uc77c \\ubd84\\uc11d \\uc911...','\\ud14c\\uc774\\ube14 \\ud30c\\uc2f1 \\uc911...','Excel \\uc2dc\\ud2b8 \\uc0dd\\uc131 \\uc911...'];
var S1A=['DSD \\ubd84\\uc11d \\uc911...','Gemini AI \\ubd84\\ub958 \\uc911... (15~30\\ucd08 \\uc18c\\uc694)','AI \\uc2dc\\ud2b8\\uba85 \\uc801\\uc6a9 \\uc911...'];
var S2=['\\ub9e4\\ud551 \\uad6c\\uc131 \\uc911...','XML \\ud328\\uce58 \\uc801\\uc6a9 \\uc911...','DSD \\uc0dd\\uc131 \\uc911...'];
var S3=['Excel \\ub370\\uc774\\ud130 \\ucd94\\ucd9c \\uc911...','Python \\uc218\\ud559 \\uac80\\uc0ac \\uc911...','Gemini AI \\uac80\\uc99d \\uc911... (30~60\\ucd08)','\\uacb0\\uacfc \\uc2dc\\ud2b8 \\uc0dd\\uc131 \\uc911...'];
var SP=['DSD XML \\ud30c\\uc2f1 \\uc911...','\\ub2f9\\uae30/\\uc804\\uae30 \\ub370\\uc774\\ud130 \\ucd94\\ucd9c \\uc911...','Gemini AI \\uc694\\uc57d \\uc911... (10~20\\ucd08)','\\ucf54\\ub4dc \\uc0dd\\uc131 \\uc911...'];
var SD=['DSD XML \\ud30c\\uc2f1 \\uc911...','\\uc140 \\ub2e8\\uc704 \\ube44\\uad50 \\uc911...','Gemini AI \\uc694\\uc57d \\uc911...','\\ub9ac\\ud3ec\\ud2b8 \\uc0dd\\uc131 \\uc911...'];
function anim(n,steps,isAI){var i=0;return setInterval(function(){if(i<steps.length)document.getElementById('pt'+n).textContent=steps[i++];},isAI?5000:1200);}
async function run1(){
  if(!F.f1)return;
  document.getElementById('btn1').disabled=true;
  var useAI=document.getElementById('chkAI').checked;
  var doRoll=document.getElementById('chkRoll').checked;
  var doNote=document.getElementById('chkNote').checked;
  var doPeriod=document.getElementById('chkPeriod').checked;
  var key=getKey();
  if(useAI&&!key){ser(1,'AI \ubd84\ub958\ub97c \uc0ac\uc6a9\ud558\ub824\uba74 Gemini API Key\ub97c \uc785\ub825\ud574\uc8fc\uc138\uc694.');document.getElementById('btn1').disabled=false;return;}
  if(doPeriod){
    var cp=parseInt(document.getElementById('curPeriod').value||'0');
    var cy=parseInt(document.getElementById('curYear').value||'0');
    if(!cp||!cy){ser(1,'\uae30\uc218\uc640 \uc5f0\ub3c4\ub97c \uc785\ub825\ud574\uc8fc\uc138\uc694.');document.getElementById('btn1').disabled=false;return;}
  }
  var isAny=useAI||doNote||doPeriod;
  sp(1,isAny?S1A[0]:S1[0],isAny);var iv=anim(1,isAny?S1A:S1,isAny);
  try{
    var fd=new FormData();
    fd.append('dsd',F.f1);
    fd.append('ai_classify',useAI?'1':'0');
    fd.append('rollover',doRoll?'1':'0');
    fd.append('note_classify',doNote?'1':'0');
    fd.append('period_change',doPeriod?'1':'0');
    if(doPeriod){
      fd.append('cur_period',document.getElementById('curPeriod').value);
      fd.append('cur_year',document.getElementById('curYear').value);
      fd.append('start_m',document.getElementById('startM').value||'1');
      fd.append('start_d',document.getElementById('startD').value||'1');
      fd.append('end_m',document.getElementById('endM').value||'12');
      fd.append('end_d',document.getElementById('endD').value||'31');
    }
    fd.append('api_key',key);fd.append('model',getModel());
    var r=await fetch('/api/dsd2excel',{method:'POST',body:fd});clearInterval(iv);ep(1);
    if(!r.ok){var e=await r.json();throw new Error(e.error||'\ubcc0\ud658 \uc2e4\ud328');}
    var blob=await r.blob();var info=JSON.parse(r.headers.get('X-Info')||'{}');
    var fname=F.f1.name.replace(/[.]dsd$/i,'')+'.xlsx';
    var sub='\uc2dc\ud2b8 '+info.sheets+'\uac1c \xb7 \uc218\uc815\uac00\ub2a5 \uc140 '+info.cells+'\uac1c \xb7 \uc7ac\ubb34\ud45c '+info.fin+'\uac1c';
    if(doRoll)sub+=' \xb7 &#x1F504;\ub864\ub85c\ubc84';
    if(doNote)sub+=' \xb7 &#x1F4DA;\uc8fc\uc11d\ubd84\ub958';
    if(doPeriod)sub+=' \xb7 &#x1F4C5;\uae30\uc218\ubcc0\uacbd';
    if(useAI)sub+=' \xb7 &#x1F916;AI\ubd84\ub958';
    sok(1,'\ubcc0\ud658 \uc644\ub8cc! Excel \ud30c\uc77c\uc744 \ub2e4\uc6b4\ub85c\ub4dc\ud558\uc138\uc694',sub,blob,fname);
  }catch(e){clearInterval(iv);ep(1);ser(1,e.message);}
  document.getElementById('btn1').disabled=false;
}
async function run2(){
  if(!F.f2||!F.f3)return;
  document.getElementById('btn2').disabled=true;
  sp(2,S2[0]);var iv=anim(2,S2);
  try{
    var fd=new FormData();fd.append('orig_dsd',F.f2);fd.append('xlsx',F.f3);
    var r=await fetch('/api/excel2dsd',{method:'POST',body:fd});clearInterval(iv);ep(2);
    if(!r.ok){var e=await r.json();throw new Error(e.error||'\\ubcc0\\ud658 \\uc2e4\\ud328');}
    var blob=await r.blob();var info=JSON.parse(r.headers.get('X-Info')||'{}');
    var fname=F.f2.name.replace(/\\.dsd$/i,'')+'_\\uc218\\uc815.dsd';
    sok(2,'DSD \\ud30c\\uc77c \\uc0dd\\uc131 \\uc644\\ub8cc!',info.tables+'\\uac1c \\ud14c\\uc774\\ube14 \\xb7 '+info.cells+'\\uac1c \\uc140 \\uc218\\uc815 \\xb7 XML '+(info.xml_ok?'\\u2713 \\uc815\\uc0c1':'\\u2717 \\uc624\\ub958'),blob,fname);
  }catch(e){clearInterval(iv);ep(2);ser(2,e.message);}
  document.getElementById('btn2').disabled=false;
}
async function run3(){
  if(!F.f4)return;
  var key=getKey();
  document.getElementById('btn3').disabled=true;
  sp(3,S3[0],true);var iv=anim(3,S3,true);
  try{
    var doNoteMap=document.getElementById('chkNoteMap').checked;
    var fd=new FormData();fd.append('xlsx',F.f4);fd.append('api_key',key);fd.append('model',getModel());
    fd.append('check_note_map',doNoteMap?'1':'0');
    var r=await fetch('/api/verify_excel',{method:'POST',body:fd});clearInterval(iv);ep(3);
    if(!r.ok){var e=await r.json();throw new Error(e.error||'\\uac80\\uc99d \\uc2e4\\ud328');}
    var blob=await r.blob();var info=JSON.parse(r.headers.get('X-Info')||'{}');
    var fname=F.f4.name.replace(/\\.xlsx$/i,'')+'_AI\\uac80\\uc99d.xlsx';
    if(info.py_result)showPy(info.py_result);
    var sub='\\uc7ac\\ubb34: '+info.fin_sheets+'\\uac1c \\xb7 \\uc8fc\\uc11d: '+info.note_sheets+'\\uac1c';
    var pe=(info.py_result&&info.py_result.errors)?info.py_result.errors.length:0;
    var pw2=(info.py_result&&info.py_result.warnings)?info.py_result.warnings.length:0;
    if(pe)sub+=' \\xb7 \\u274C\\uc624\\ub958 '+pe+'\\uac74';
    if(pw2)sub+=' \\xb7 \\u26A0\\uFE0F\\uacbd\\uace0 '+pw2+'\\uac74';
    if(!key)sub+=' \\xb7 (Gemini \\uac80\\uc99d \\uc0dd\\ub7b5)';
    sok(3,'AI \\uac80\\uc99d \\uc644\\ub8cc!',sub,blob,fname);
    if(info.preview){document.getElementById('vtext').textContent=info.preview;document.getElementById('vbox').style.display='block';}
  }catch(e){clearInterval(iv);ep(3);ser(3,e.message);}
  document.getElementById('btn3').disabled=false;
}
async function runPrior(){
  if(!F.f5||!F.f6)return;
  document.getElementById('btn4').disabled=true;
  sp(4,SP[0],true);var iv=anim(4,SP,true);
  try{
    var fd=new FormData();fd.append('prev_dsd',F.f5);fd.append('curr_dsd',F.f6);fd.append('api_key',getKey());fd.append('model',getModel());
    var r=await fetch('/api/validate_prior',{method:'POST',body:fd});clearInterval(iv);ep(4);
    if(!r.ok){var e=await r.json();throw new Error(e.error||'\\uac80\\uc99d \\uc2e4\\ud328');}
    var blob=await r.blob();var info=JSON.parse(r.headers.get('X-Info')||'{}');
    var fname='\\uc804\\uae30\\uae08\\uc561\\uac80\\uc99d_'+new Date().toISOString().slice(0,10)+'.xlsx';
    var warn=info.mismatches>0?(' \\xb7 \\u274C\\uC804\\uAE30\\uC774\\uC6D4\\uC624\\uB958 '+info.mismatches+'\\uAC74'):'';
    sok(4,'\\uc804\\uae30\\uae08\\uc561 \\uac80\\uc99d \\uc644\\ub8cc!','\\ub9e4\\ud551 \\ud56d\\ubaa9: '+info.total+'\\uac1c \\xb7 \\uc77c\\uce58: '+info.matches+'\\uac74'+warn,blob,fname);
  }catch(e){clearInterval(iv);ep(4);ser(4,e.message);}
  document.getElementById('btn4').disabled=false;
}
async function runDiff(){
  if(!F.f7||!F.f8)return;
  document.getElementById('btn5').disabled=true;
  sp(5,SD[0],true);var iv=anim(5,SD,true);
  try{
    var fd=new FormData();fd.append('dsd_a',F.f7);fd.append('dsd_b',F.f8);fd.append('api_key',getKey());fd.append('model',getModel());
    var r=await fetch('/api/compare_dsd',{method:'POST',body:fd});clearInterval(iv);ep(5);
    if(!r.ok){var e=await r.json();throw new Error(e.error||'\\ube44\\uad50 \\uc2e4\\ud328');}
    var blob=await r.blob();var info=JSON.parse(r.headers.get('X-Info')||'{}');
    var fname='DSD_\\ube44\\uad50_'+new Date().toISOString().slice(0,10)+'.xlsx';
    sok(5,'\\ube44\\uad50 \\ubd84\\uc11d \\uc644\\ub8cc!','EXTRACTION \\ubcc0\\uacbd: '+info.ext_diffs+'\\uac74 \\xb7 \\uc140 \\ubcc0\\uacbd: '+info.cell_diffs+'\\uac74',blob,fname);
  }catch(e){clearInterval(iv);ep(5);ser(5,e.message);}
  document.getElementById('btn5').disabled=false;
}
function showKill(){document.getElementById('killModal').classList.add('show');}
function hideKill(){document.getElementById('killModal').classList.remove('show');}
async function doKill(){
  hideKill();
  localStorage.removeItem('easydsd_k');
  localStorage.removeItem('easydsd_m');
  try{await fetch('/api/shutdown',{method:'POST'});}catch(e){}
  document.body.innerHTML='<div style="display:flex;align-items:center;justify-content:center;height:100vh;font-family:sans-serif;color:#556;font-size:15px;">\\uc11c\\ubc84\\uac00 \\uc885\\ub8cc\\ub418\\uc5c8\\uc2b5\\ub2c8\\ub2e4. \\uc774 \\ud0ed\\uc744 \\ub2eb\\uc544\\uc8fc\\uc138\\uc694.</div>';
}
</script>
</body>
</html>"""



# ── Flask 라우트 ──────────────────────────────────────────────────────────────
@app.route('/')
def index(): return Response(HTML, mimetype='text/html; charset=utf-8')

@app.route('/api/heartbeat', methods=['POST'])
def api_heartbeat():
    global _last_ping; _last_ping=time.time(); return jsonify(ok=True)

@app.route('/api/shutdown', methods=['POST'])
def api_shutdown():
    threading.Thread(target=lambda:(time.sleep(0.3),os._exit(0)),daemon=True).start()
    return jsonify(ok=True)

@app.route('/api/dsd2excel', methods=['POST'])
def api_dsd2excel():
    try:
        dsd_bytes        = request.files['dsd'].read()
        ai_classify      = request.form.get('ai_classify','0')=='1'
        do_rollover      = request.form.get('rollover','0')=='1'
        do_note_classify = request.form.get('note_classify','0')=='1'
        do_period_change = request.form.get('period_change','0')=='1'
        api_key          = request.form.get('api_key','').strip()
        model_name       = request.form.get('model','gemini-3-flash-preview').strip()
        period_params=None
        if do_period_change:
            try:
                period_params={
                    'cur_period':int(request.form.get('cur_period','1')),
                    'cur_year':  int(request.form.get('cur_year','2026')),
                    'start_m':   int(request.form.get('start_m','1')),
                    'start_d':   int(request.form.get('start_d','1')),
                    'end_m':     int(request.form.get('end_m','12')),
                    'end_d':     int(request.form.get('end_d','31')),
                }
            except ValueError:
                return jsonify(error='기수/연도 입력값을 확인해주세요.'),400
        ai_mapping={}
        if ai_classify:
            if not api_key:
                return jsonify(error='AI 분류를 사용하려면 Gemini API Key를 입력해주세요.'),400
            xml_raw=zipfile.ZipFile(io.BytesIO(dsd_bytes)).read('contents.xml').decode('utf-8',errors='replace')
            _,tables=parse_xml(xml_raw)
            if not do_note_classify:
                ai_mapping=gemini_classify_tables(api_key,tables,model_name)
        r_key=api_key if (do_rollover or do_note_classify) else ''
        xlsx=dsd_to_excel_bytes(dsd_bytes,ai_mapping or None,
                                do_rollover=do_rollover,
                                rollover_api_key=r_key,
                                rollover_model=model_name,
                                do_note_classify=do_note_classify,
                                do_period_change=do_period_change,
                                period_params=period_params)
        wb=openpyxl.load_workbook(io.BytesIO(xlsx),data_only=True)
        cells=sum(1 for ws in wb.worksheets for row in ws.iter_rows()
                  for cell in row if cell.fill and cell.fill.fill_type=='solid'
                  and cell.fill.fgColor and cell.fill.fgColor.type=='rgb'
                  and cell.fill.fgColor.rgb.upper().endswith(EDIT_COLOR.upper()))
        fin=[ws.title for ws in wb.worksheets if any(ws.title.startswith(e) for e in FIN_PREFIXES)]
        resp=send_file(io.BytesIO(xlsx),
                       mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       as_attachment=True,download_name='converted.xlsx')
        resp.headers['X-Info']=json.dumps(
            {'sheets':len(wb.sheetnames),'cells':cells,'fin':len(fin),'ai':bool(ai_mapping),'rollover':do_rollover})
        return resp
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify(error=str(e)),500

@app.route('/api/excel2dsd', methods=['POST'])
def api_excel2dsd():
    try:
        orig=request.files['orig_dsd'].read(); xlsx=request.files['xlsx'].read()
        dsd=excel_to_dsd_bytes(orig,xlsx)
        import xml.etree.ElementTree as ET
        with zipfile.ZipFile(io.BytesIO(dsd)) as z: xt=z.read('contents.xml').decode('utf-8')
        xml_ok=True
        try: ET.fromstring(xt)
        except: xml_ok=False
        wb=openpyxl.load_workbook(io.BytesIO(xlsx),data_only=True)
        tc=tb=0
        for sname in wb.sheetnames:
            if sname in ('📋사용안내','_원본XML'): continue
            ws=wb[sname]
            cnt=sum(1 for row in ws.iter_rows() for cell in row
                    if cell.fill and cell.fill.fill_type=='solid'
                    and cell.fill.fgColor and cell.fill.fgColor.type=='rgb'
                    and cell.fill.fgColor.rgb.upper().endswith(EDIT_COLOR.upper()))
            if cnt: tb+=1; tc+=cnt
        resp=send_file(io.BytesIO(dsd),mimetype='application/octet-stream',
                       as_attachment=True,download_name='output.dsd')
        resp.headers['X-Info']=json.dumps({'tables':tb,'cells':tc,'xml_ok':xml_ok})
        return resp
    except Exception as e:
        return jsonify(error=str(e)),500

@app.route('/api/verify_excel', methods=['POST'])
def api_verify_excel():
    try:
        xlsx_bytes=request.files['xlsx'].read()
        api_key=request.form.get('api_key','').strip()
        model_name=request.form.get('model','gemini-3-flash-preview').strip()
        check_note_map=request.form.get('check_note_map','0')=='1'
        py_result=python_verify(xlsx_bytes)
        fin_data,note_data=extract_fin_and_notes(xlsx_bytes)
        if not fin_data:
            return jsonify(error='재무제표 시트(🏦💹📈💰)를 찾을 수 없습니다.'),400
        note_map_result=py_result.get('note_map') if check_note_map else None
        if api_key:
            verify_result=gemini_verify_enhanced(api_key,fin_data,note_data,py_result,model_name,
                                                  note_map_result=note_map_result)
        else:
            lines=([f'[오류] {e}' for e in py_result.get('errors',[])]
                   +[f'[경고] {w}' for w in py_result.get('warnings',[])]
                   +[f'[정보] {i}' for i in py_result.get('info',[])])
            verify_result=(
                '## ✅ 파이썬 수학 검사 결과\n'+('\n'.join(lines) or '이상 없음')
                +'\n\n## 📋 종합 의견\nGemini API Key가 없어 AI 교차 검증은 생략되었습니다.'
            )
        wb=openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        if '🤖AI검증결과' in wb.sheetnames: del wb['🤖AI검증결과']
        ws_v=wb.create_sheet('🤖AI검증결과',0)
        ws_v.sheet_view.showGridLines=False
        tc=ws_v.cell(1,1,'🤖 Gemini AI + Python 재무제표 검증 결과 (easydsd v0.02)')
        tc.fill=PatternFill('solid',fgColor='4A148C'); tc.font=Font(color='FFFFFF',bold=True,size=12)
        tc.alignment=Alignment(horizontal='left',vertical='center')
        ws_v.merge_cells('A1:F1'); ws_v.row_dimensions[1].height=28
        sc=ws_v.cell(2,1,
            f'생성: {time.strftime("%Y-%m-%d %H:%M")}  |  재무: {len(fin_data)}개  |  주석: {len(note_data)}개  |'
            f'  Python오류: {len(py_result.get("errors",[]))}건  경고: {len(py_result.get("warnings",[]))}건')
        sc.font=Font(color='7B1FA2',size=9,italic=True); ws_v.row_dimensions[2].height=16
        CM={'## ✅':('E8F5E9','1B5E20'),'## ❌':('FFEBEE','B71C1C'),'## ⚠':('FFF8E1','E65100'),'## 📋':('E3F2FD','0D47A1')}
        for ri,line in enumerate(verify_result.split('\n'),4):
            cell=ws_v.cell(ri,1,line)
            matched=next(((fg,fc) for k,(fg,fc) in CM.items() if line.startswith(k)),None)
            if matched:
                cell.fill=PatternFill('solid',fgColor=matched[0]); cell.font=Font(bold=True,size=10,color=matched[1])
            else:
                cell.font=Font(size=9)
            cell.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
            ws_v.row_dimensions[ri].height=16
        ws_v.column_dimensions['A'].width=90
        buf=io.BytesIO(); wb.save(buf)
        preview=verify_result[:600]+('...' if len(verify_result)>600 else '')
        resp=send_file(io.BytesIO(buf.getvalue()),
                       mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       as_attachment=True,download_name='verified.xlsx')
        resp.headers['X-Info']=json.dumps(
            {'fin_sheets':len(fin_data),'note_sheets':len(note_data),
             'preview':preview,'py_result':py_result,
             'note_map':py_result.get('note_map',{})})
        return resp
    except Exception as e:
        return jsonify(error=str(e)),500

@app.route('/api/validate_prior', methods=['POST'])
def api_validate_prior():
    """전기금액 검증: prev_dsd 당기 == curr_dsd 전기"""
    try:
        prev_dsd  = request.files['prev_dsd'].read()
        curr_dsd  = request.files['curr_dsd'].read()
        api_key   = request.form.get('api_key','').strip()
        model_name= request.form.get('model','gemini-3-flash-preview').strip()

        result=validate_prior_period(prev_dsd,curr_dsd,api_key,model_name)
        mismatches=result['mismatches']; matches=result['matches']
        ai_report =result['ai_report'];  curr_data=result['curr_data']

        # Excel 생성
        wb=openpyxl.Workbook()

        # 시트1: AI 리포트
        ws0=wb.active; ws0.title='📋AI비교검증리포트'; ws0.sheet_view.showGridLines=False
        tc=ws0.cell(1,1,'📋 전기금액 검증 리포트 (수정전 당기 vs 수정후 전기)')
        tc.fill=PatternFill('solid',fgColor='00695C'); tc.font=Font(color='FFFFFF',bold=True,size=13)
        tc.alignment=Alignment(horizontal='left',vertical='center')
        ws0.merge_cells('A1:E1'); ws0.row_dimensions[1].height=30
        sc=ws0.cell(2,1,
            f'생성: {time.strftime("%Y-%m-%d %H:%M")}  |  비교 항목: {len(matches)+len(mismatches)}건  |  '
            f'일치: {len(matches)}건  |  전기이월 오류: {len(mismatches)}건')
        sc.font=Font(color='004D40',size=9,italic=True); ws0.row_dimensions[2].height=15
        CM2={'## ⚠️':('FFF8E1','E65100'),'## 📊':('E0F2F1','004D40'),
             '## ✅':('E8F5E9','1B5E20'),'## 📋':('E3F2FD','0D47A1')}
        for ri,line in enumerate(ai_report.split('\n'),4):
            c=ws0.cell(ri,1,line)
            matched=next(((fg,fc) for k,(fg,fc) in CM2.items() if line.startswith(k)),None)
            if matched:
                c.fill=PatternFill('solid',fgColor=matched[0]); c.font=Font(bold=True,size=10,color=matched[1])
            else:
                c.font=Font(size=9)
            c.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
            ws0.row_dimensions[ri].height=16
        ws0.column_dimensions['A'].width=100

        # 시트2: 전기이월 오류 목록
        if mismatches:
            ws_mm=wb.create_sheet('❌전기이월오류')
            for ci,(h,w) in enumerate([('재무제표',18),('계정과목',35),('수정전당기',18),('수정후전기',18),('차이',18)],1):
                c=ws_mm.cell(1,ci,h)
                c.fill=PatternFill('solid',fgColor='B71C1C'); c.font=Font(color='FFFFFF',bold=True,size=9)
                c.alignment=Alignment(horizontal='center',vertical='center')
                ws_mm.column_dimensions[get_column_letter(ci)].width=w
            for ri,(label,acct,prev,curr,diff) in enumerate(mismatches,2):
                ws_mm.cell(ri,1,label).font=Font(bold=True,size=9)
                ws_mm.cell(ri,2,acct).font=Font(size=9)
                if prev is not None:
                    c3=ws_mm.cell(ri,3,f'{prev:,.0f}'); c3.font=Font(size=9); c3.alignment=Alignment(horizontal='right')
                if curr is not None:
                    c4=ws_mm.cell(ri,4,f'{curr:,.0f}'); c4.font=Font(size=9); c4.alignment=Alignment(horizontal='right')
                if diff is not None:
                    c5=ws_mm.cell(ri,5,f'{diff:,.0f}')
                    c5.fill=PatternFill('solid',fgColor='FFCDD2'); c5.font=Font(size=9,bold=True,color='B71C1C')
                    c5.alignment=Alignment(horizontal='right')
                # 오류 행 빨간 하이라이트
                for ci2 in range(1,6):
                    ws_mm.cell(ri,ci2).fill=PatternFill('solid',fgColor='FFEBEE')
                ws_mm.cell(ri,2).fill=PatternFill('solid',fgColor='FF0000')
                ws_mm.cell(ri,2).font=Font(size=9,color='FFFFFF',bold=True)

        # 시트3: 수정후 전체 재무 데이터
        ws_all=wb.create_sheet('📊수정후재무데이터')
        for ci,(h,w) in enumerate([('재무제표',18),('계정과목',35),('당기',18),('전기',18),('상태',10)],1):
            c=ws_all.cell(1,ci,h)
            c.fill=PatternFill('solid',fgColor='1F4E79'); c.font=Font(color='FFFFFF',bold=True,size=9)
            c.alignment=Alignment(horizontal='center',vertical='center')
            ws_all.column_dimensions[get_column_letter(ci)].width=w
        # mismatches로 오류 계정 set 만들기
        err_keys={(lb,ac) for lb,ac,_p,_c,_d in mismatches}
        for ri,(label,acct,cur,pri) in enumerate(curr_data,2):
            ws_all.cell(ri,1,label).font=Font(size=9)
            ws_all.cell(ri,2,acct).font=Font(size=9)
            if cur is not None:
                c3=ws_all.cell(ri,3,f'{cur:,.0f}'); c3.font=Font(size=9); c3.alignment=Alignment(horizontal='right')
            if pri is not None:
                c4=ws_all.cell(ri,4,f'{pri:,.0f}'); c4.font=Font(size=9); c4.alignment=Alignment(horizontal='right')
            if (label,acct) in err_keys:
                for ci2 in range(1,5):
                    ws_all.cell(ri,ci2).fill=PatternFill('solid',fgColor='FFEBEE')
                ws_all.cell(ri,5,'❌오류').font=Font(size=9,bold=True,color='B71C1C')
            else:
                ws_all.cell(ri,5,'✓').font=Font(size=9,color='1B5E20')

        buf=io.BytesIO(); wb.save(buf)
        resp=send_file(io.BytesIO(buf.getvalue()),
                       mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       as_attachment=True,download_name='prior_validation.xlsx')
        resp.headers['X-Info']=json.dumps({
            'mismatches':len(mismatches),
            'matches':len(matches),
            'total':len(mismatches)+len(matches)
        })
        return resp
    except Exception as e:
        return jsonify(error=str(e)),500

@app.route('/api/compare_dsd', methods=['POST'])
def api_compare_dsd():
    try:
        dsd_a=request.files['dsd_a'].read(); dsd_b=request.files['dsd_b'].read()
        api_key=request.form.get('api_key','').strip()
        model_name=request.form.get('model','gemini-3-flash-preview').strip()
        result=compare_dsd_bytes(dsd_a,dsd_b,api_key,model_name)
        da=parse_dsd_tables(dsd_a); db=parse_dsd_tables(dsd_b)
        ext_diffs=sum(1 for k in set(list(da['exts'].keys())+list(db['exts'].keys()))
                      if da['exts'].get(k)!=db['exts'].get(k))
        cell_diffs=sum(1 for ti in sorted(set(list(da['tables'].keys())+list(db['tables'].keys())))
                       for ri in range(max(len(da['tables'].get(ti,[])),len(db['tables'].get(ti,[]))))
                       for ci in range(max(len(da['tables'].get(ti,[])[ri] if ri<len(da['tables'].get(ti,[])) else []),
                                           len(db['tables'].get(ti,[])[ri] if ri<len(db['tables'].get(ti,[])) else [])))
                       if (da['tables'].get(ti,[])[ri][ci] if ri<len(da['tables'].get(ti,[])) and ci<len(da['tables'].get(ti,[])[ri]) else '')
                          != (db['tables'].get(ti,[])[ri][ci] if ri<len(db['tables'].get(ti,[])) and ci<len(db['tables'].get(ti,[])[ri]) else ''))
        resp=send_file(io.BytesIO(result),
                       mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       as_attachment=True,download_name='diff_report.xlsx')
        resp.headers['X-Info']=json.dumps({'ext_diffs':ext_diffs,'cell_diffs':cell_diffs})
        return resp
    except Exception as e:
        return jsonify(error=str(e)),500

# ── 실행 ─────────────────────────────────────────────────────────────────────
def open_browser():
    time.sleep(1.5)
    webbrowser.open(f'http://127.0.0.1:{PORT}')

if __name__=='__main__':
    print('='*54)
    print('  easydsd v0.02 - DART 감사보고서 변환 + AI')
    print(f'  http://127.0.0.1:{PORT}')
    print('  종료: 브라우저 종료 버튼 or Ctrl+C')
    print('='*54)
    threading.Thread(target=open_browser,daemon=True).start()
    app.run(host='127.0.0.1',port=PORT,debug=False,threaded=True)
